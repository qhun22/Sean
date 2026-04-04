"""
Admin Views (dashboard, products, brands, users, banners) - QHUN22 Store
Auto-generated from views.py
"""
import os
import json
import uuid
import random
import time
import re
import traceback
import requests
from decimal import Decimal
from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.conf import settings
from django.http import JsonResponse
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods, require_POST
from django.db import models
from django.db.models import Q, Count, Sum, Max
from django.utils import timezone
from django.urls import reverse
import datetime



@login_required
def dashboard_view(request):
    """
    Trang dashboard quản trị - hiển thị danh sách người dùng
    Chỉ admin (superuser) mới có thể truy cập
    """
    # Kiểm tra quyền admin
    if not request.user.is_superuser:
        messages.error(request, 'Bạn không có quyền truy cập trang này!')
        return redirect('store:profile')
    
    # Lấy danh sách tất cả người dùng
    from store.models import CustomUser, SiteVisit
    all_users = CustomUser.objects.all()
    users = all_users.order_by('-date_joined')
    
    # Search users
    user_search = request.GET.get('user_search', '')
    if user_search:
        users = users.filter(
            models.Q(email__icontains=user_search) | 
            models.Q(last_name__icontains=user_search) |
            models.Q(first_name__icontains=user_search) |
            models.Q(phone__icontains=user_search)
        )
    
    # Phân trang người dùng (15 người dùng/trang)
    user_page = request.GET.get('user_page', 1)
    user_paginator = Paginator(users, 15)
    try:
        users_paginated = user_paginator.page(user_page)
    except PageNotAnInteger:
        users_paginated = user_paginator.page(1)
    except EmptyPage:
        users_paginated = user_paginator.page(user_paginator.num_pages)
    
    # Thống kê
    regular_users = all_users.filter(is_oauth_user=False, is_superuser=False).count()
    oauth_users = all_users.filter(is_oauth_user=True).count()
    admin_users = all_users.filter(is_superuser=True).count()
    filtered_users_count = users.count()
    
    # Tổng lượt truy cập trang chủ
    total_visits = SiteVisit.objects.count()
    
    # Sản phẩm sắp hết hàng (dưới 5 sản phẩm)
    from store.models import Product, Order, Brand
    low_stock_products_count = Product.objects.filter(stock__gt=0, stock__lt=5).count()
    
    # Danh sách hãng với số sản phẩm (cho stats)
    brands_for_stats = Brand.objects.filter(is_active=True).annotate(product_count=Count('products')).order_by('name')[:8]
    # Danh sách hãng đầy đủ (cho dropdown/select ở dashboard)
    brands_all_active = Brand.objects.filter(is_active=True).order_by('name')
    
    # Danh sách hãng cho bảng (có phân trang)
    all_brands_base = Brand.objects.filter(is_active=True).annotate(product_count=Count('products')).order_by('name')
    all_brands = all_brands_base
    brand_search = request.GET.get('brand_search', '')
    if brand_search:
        all_brands = all_brands.filter(name__icontains=brand_search)

    brand_total_count = all_brands_base.count()
    brands_with_products_count = all_brands_base.filter(product_count__gt=0).count()
    brands_without_products_count = all_brands_base.filter(product_count=0).count()
    brand_filtered_count = all_brands.count()
    
    brand_page = request.GET.get('brand_page', 1)
    brand_paginator = Paginator(all_brands, 15)
    try:
        brands_paginated = brand_paginator.page(brand_page)
    except PageNotAnInteger:
        brands_paginated = brand_paginator.page(1)
    except EmptyPage:
        brands_paginated = brand_paginator.page(brand_paginator.num_pages)
    
    # Danh sách sản phẩm (Product)
    from store.models import Product
    all_products_base = Product.objects.select_related('brand', 'detail').all().order_by('-created_at')
    all_products = all_products_base
    product_search = request.GET.get('product_search', '')
    if product_search:
        all_products = all_products.filter(name__icontains=product_search)

    discounted_products_count = sum(1 for product in all_products_base if getattr(getattr(product, 'detail', None), 'summary_discount_percent', 0) > 0)
    product_filtered_count = all_products.count()
    
    product_page = request.GET.get('product_page', 1)
    product_paginator = Paginator(all_products, 15)
    try:
        products_paginated = product_paginator.page(product_page)
    except PageNotAnInteger:
        products_paginated = product_paginator.page(1)
    except EmptyPage:
        products_paginated = product_paginator.page(product_paginator.num_pages)
    
    # Doanh thu hôm nay - chỉ tính khi đơn hàng đã giao thành công (delivered)
    today = timezone.now().date()
    revenue_today = Order.objects.filter(
        created_at__date=today,
        status='delivered'
    ).aggregate(total=Sum('total_amount'))['total'] or 0

    # Thống kê đơn hàng
    from datetime import datetime, timedelta
    now = timezone.now()
    current_year = now.year
    current_month = now.month

    orders_today_count = Order.objects.filter(created_at__date=today).count()
    orders_this_month_count = Order.objects.filter(created_at__year=current_year, created_at__month=current_month).count()
    orders_this_year_count = Order.objects.filter(created_at__year=current_year).count()
    orders_pending_count = Order.objects.filter(status='pending').count()
    orders_processing_count = Order.objects.filter(status='processing').count()
    orders_shipped_count = Order.objects.filter(status='shipped').count()
    orders_delivered_count = Order.objects.filter(status='delivered').count()
    # W1 FIX: "Đã hủy" chỉ đếm đơn bị hủy thật, không có yêu cầu hoàn tiền
    orders_cancelled_count = Order.objects.filter(status='cancelled', refund_status='').count()
    orders_payment_expired_count = Order.objects.filter(status='payment_expired').count()
    # Đơn cần hoàn tiền: có refund_status='pending' (bất kể status đơn)
    orders_refund_pending_count = Order.objects.filter(refund_status='pending').count()

    # VietQR fail rate (expired + cancelled payments)
    vietqr_total = Order.objects.filter(payment_method='vietqr').count()
    vietqr_failed = Order.objects.filter(payment_method='vietqr', payment_status__in=['cancelled', 'expired']).count()
    vietqr_fail_rate = round(vietqr_failed / vietqr_total * 100, 1) if vietqr_total > 0 else 0

    total_orders_all = Order.objects.count()
    total_revenue_all = Order.objects.filter(status='delivered').aggregate(t=Sum('total_amount'))['t'] or 0
    avg_order_value = int(total_revenue_all / total_orders_all) if total_orders_all > 0 else 0
    conversion_rate = round(total_orders_all / total_visits * 100, 2) if total_visits > 0 else 0

    # Doanh thu tháng này - chỉ tính khi đơn hàng đã giao thành công (delivered)
    # Sử dụng __month và __year để tránh vấn đề timezone
    revenue_this_month = Order.objects.filter(
        created_at__year=current_year,
        created_at__month=current_month,
        status='delivered'
    ).aggregate(total=Sum('total_amount'))['total'] or 0

    # Doanh thu hôm qua & 7 ngày gần nhất
    yesterday = today - timedelta(days=1)
    revenue_yesterday = Order.objects.filter(
        created_at__date=yesterday,
        status='delivered'
    ).aggregate(total=Sum('total_amount'))['total'] or 0

    seven_days_ago = today - timedelta(days=6)
    revenue_7days = Order.objects.filter(
        created_at__date__gte=seven_days_ago,
        status='delivered'
    ).aggregate(total=Sum('total_amount'))['total'] or 0
    
    # Dữ liệu biểu đồ doanh thu năm hiện tại (tự động theo năm)
    from django.db.models.functions import ExtractMonth
    
    # Lấy năm hiện tại từ timezone
    current_year = now.year
    
    # Lấy doanh thu từng tháng trong năm hiện tại - chỉ tính đơn hàng đã giao
    monthly_revenue = {}
    for month in range(1, 13):
        monthly_revenue[month] = 0
    
    # Query orders with delivered status only
    orders_current_year = Order.objects.filter(
        created_at__year=current_year,
        status='delivered'
    )
    
    for order in orders_current_year:
        month = order.created_at.month
        monthly_revenue[month] += float(order.total_amount)
    
    # Tính toán chiều cao biểu đồ
    max_revenue = max(monthly_revenue.values()) if max(monthly_revenue.values()) > 0 else 1
    
    month_data = []
    month_names = ['T1', 'T2', 'T3', 'T4', 'T5', 'T6', 'T7', 'T8', 'T9', 'T10', 'T11', 'T12']
    for month in range(1, 13):
        value = monthly_revenue[month]
        # Chiều cao tối đa 280px, tối thiểu 8px
        height = int((value / max_revenue) * 280) if value > 0 else 8
        month_data.append({
            'name': month_names[month - 1],
            'value': int(value),
            'height': height,
        })
    
    # Tổng chi phí nhập hàng (giá vốn × số lượng tồn)
    all_products_for_cost = Product.objects.only('id', 'name', 'cost_price', 'stock').order_by('name')
    total_import_cost = 0
    has_cost_data = False
    cost_price_products = []
    for p in all_products_for_cost:
        cp = p.cost_price
        if cp is not None:
            has_cost_data = True
            total_import_cost += int(cp) * p.stock
        cost_price_products.append({
            'id': p.id,
            'name': p.name,
            'stock': p.stock,
            'cost_price': int(cp) if cp is not None else None,
            'subtotal': int(cp) * p.stock if cp is not None else None,
        })

    context = {
        'users_paginated': users_paginated,
        'total_users': all_users.count(),
        'regular_users': regular_users,
        'oauth_users': oauth_users,
        'admin_users': admin_users,
        'filtered_users_count': filtered_users_count,
        'total_visits': total_visits,
        'low_stock_products_count': low_stock_products_count,
        'revenue_today': revenue_today,
        'revenue_yesterday': revenue_yesterday,
        'revenue_7days': revenue_7days,
        'revenue_this_month': revenue_this_month,
        'months': month_data,
        'chart_year': current_year,
        # NOTE: dashboard.html dùng `brands` cho các dropdown (chọn hãng) => phải là danh sách đầy đủ
        'brands': brands_all_active,
        # nếu cần hiển thị 8 hãng cho thống kê/box, dùng biến riêng
        'brands_for_stats': brands_for_stats,
        'brands_paginated': brands_paginated,
        'brand_total_count': brand_total_count,
        'brands_with_products_count': brands_with_products_count,
        'brands_without_products_count': brands_without_products_count,
        'brand_filtered_count': brand_filtered_count,
        'products_paginated': products_paginated,
        'product_filtered_count': product_filtered_count,
        'discounted_products_count': discounted_products_count,
        'products': all_products[:50],  # For SKU dropdown
        # Order stats
        'orders_today': orders_today_count,
        'orders_this_month': orders_this_month_count,
        'orders_this_year': orders_this_year_count,
        'orders_pending': orders_pending_count,
        'orders_processing': orders_processing_count,
        'orders_shipped': orders_shipped_count,
        'orders_delivered': orders_delivered_count,
        'orders_cancelled': orders_cancelled_count,
        'orders_payment_expired': orders_payment_expired_count,
        'orders_refund_pending': orders_refund_pending_count,
        'vietqr_total': vietqr_total,
        'vietqr_failed': vietqr_failed,
        'vietqr_fail_rate': vietqr_fail_rate,
        'avg_order_value': avg_order_value,
        'conversion_rate': conversion_rate,
        # Product stats
        'total_products': Product.objects.count(),
        'active_products': Product.objects.filter(stock__gt=0).count(),
        'out_of_stock_products': Product.objects.filter(stock=0).count(),
        'total_stock': Product.objects.aggregate(t=Sum('stock'))['t'] or 0,
        # Import cost stats
        'total_import_cost': total_import_cost,
        'cost_price_products': cost_price_products,
        'has_cost_data': has_cost_data,
    }
    return render(request, 'store/admin/dashboard.html', context)


@login_required
@require_http_methods(["GET"])
def dashboard_order_detail(request):
    """API: Chi tiết đơn hàng theo filter cho dashboard stat boxes"""
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'message': 'Không có quyền!'}, status=403)

    from store.models import Order, OrderItem
    from django.db.models import Prefetch

    filter_type = request.GET.get('filter', 'today')
    page = max(1, int(request.GET.get('page', 1) or 1))
    search = request.GET.get('search', '').strip()
    status_sub = request.GET.get('status_sub', '').strip()  # lọc thêm theo status trong time-based filter
    per_page = 10

    now = timezone.now()
    today = now.date()

    TITLE_MAP = {
        'today': f'Danh sách đơn hàng hôm nay (00:00 → 23:59 {today.strftime("%d/%m/%Y")})',
        'month': f'Danh sách đơn hàng tháng {now.month}/{now.year}',
        'year': f'Danh sách đơn hàng năm {now.year}',
        'pending': 'Danh sách đơn chờ xử lý (Pending)',
        'processing': 'Danh sách đơn đang xử lý (Processing)',
        'shipped': 'Danh sách đơn đang giao (Shipped)',
        'delivered': 'Danh sách đơn giao thành công (Delivered)',
        'cancelled': 'Danh sách đơn đã hủy (Cancelled)',
        'refund_pending': 'Danh sách đơn cần hoàn tiền',
    }

    STATUS_DISPLAY = {
        'awaiting_payment': 'Chờ TT',
        'pending': 'Đã đặt',
        'processing': 'Xử lý',
        'shipped': 'Đang giao',
        'delivered': 'Đã giao',
        'cancelled': 'Đã hủy',
        'payment_expired': 'Hết hạn TT',
        'refund': 'Y/c hoàn tiền',
        'refunded': 'Đã hoàn tiền',
        'refund_pending': 'Chờ hoàn tiền',
    }

    # --- Base queryset theo filter_type ---
    base_qs = Order.objects.select_related('user').prefetch_related('items').order_by('-created_at')
    TIME_FILTERS = ('today', 'month', 'year')
    is_time_filter = filter_type in TIME_FILTERS

    if filter_type == 'today':
        base_qs = base_qs.filter(created_at__date=today)
    elif filter_type == 'month':
        base_qs = base_qs.filter(created_at__year=now.year, created_at__month=now.month)
    elif filter_type == 'year':
        base_qs = base_qs.filter(created_at__year=now.year)
    elif filter_type == 'refund_pending':
        base_qs = base_qs.filter(refund_status='pending')
    elif filter_type in ('pending', 'processing', 'shipped', 'delivered', 'cancelled', 'payment_expired'):
        base_qs = base_qs.filter(status=filter_type)

    # --- Stats từ base_qs (trước khi lọc thêm) ---
    total_base = base_qs.count()
    total_value_base = base_qs.aggregate(s=Sum('total_amount'))['s'] or 0
    total_revenue_base = base_qs.filter(status='delivered').aggregate(s=Sum('total_amount'))['s'] or 0
    # W1 FIX: tỷ lệ hủy chỉ tính đơn hủy thật
    total_cancelled_base = base_qs.filter(status='cancelled').count()

    # Tổng SP bán (chỉ đơn delivered)
    delivered_ids = base_qs.filter(status='delivered').values_list('id', flat=True)
    items_sold_base = OrderItem.objects.filter(order_id__in=delivered_ids).aggregate(s=Sum('quantity'))['s'] or 0

    # Tỷ lệ hủy %
    cancel_rate = round(total_cancelled_base / total_base * 100, 1) if total_base > 0 else 0

    # Growth so với cùng kỳ năm trước (chỉ cho filter year)
    growth_label = '-'
    if filter_type == 'year':
        last_year_count = Order.objects.filter(created_at__year=now.year - 1).count()
        if last_year_count > 0:
            growth = round((total_base - last_year_count) / last_year_count * 100, 1)
            growth_label = ('+' if growth >= 0 else '') + str(growth) + '%'
        else:
            growth_label = 'N/A (chưa có dữ liệu)'

    # Build stats strip (4 cards)
    def fmt_price(v): return f"{int(v):,}đ".replace(',', '.')
    if filter_type == 'today':
        stat_cards = [
            {'label': 'Tổng đơn hôm nay', 'value': str(total_base)},
            {'label': 'Doanh thu hôm nay', 'value': fmt_price(total_revenue_base)},
            {'label': 'Tổng SP đã bán', 'value': str(items_sold_base) + ' SP'},
            {'label': 'Đơn đã hủy', 'value': str(total_cancelled_base)},
        ]
    elif filter_type == 'month':
        stat_cards = [
            {'label': 'Tổng đơn tháng', 'value': str(total_base)},
            {'label': 'Doanh thu tháng', 'value': fmt_price(total_revenue_base)},
            {'label': 'Tổng SP đã bán', 'value': str(items_sold_base) + ' SP'},
            {'label': 'Tỷ lệ hủy', 'value': str(cancel_rate) + '%'},
        ]
    elif filter_type == 'year':
        stat_cards = [
            {'label': 'Tổng đơn năm', 'value': str(total_base)},
            {'label': 'Doanh thu năm', 'value': fmt_price(total_revenue_base)},
            {'label': 'Tổng SP đã bán', 'value': str(items_sold_base) + ' SP'},
            {'label': f'Tăng trưởng vs {now.year - 1}', 'value': growth_label},
        ]
    elif filter_type == 'pending':
        stat_cards = [
            {'label': 'Tổng đơn pending', 'value': str(total_base)},
            {'label': 'Tổng giá trị', 'value': fmt_price(total_value_base)},
            {'label': '', 'value': ''},
            {'label': '', 'value': ''},
        ]
    elif filter_type == 'processing':
        stat_cards = [
            {'label': 'Tổng đơn xử lý', 'value': str(total_base)},
            {'label': 'Tổng giá trị', 'value': fmt_price(total_value_base)},
            {'label': '', 'value': ''},
            {'label': '', 'value': ''},
        ]
    elif filter_type == 'shipped':
        stat_cards = [
            {'label': 'Tổng đơn đang giao', 'value': str(total_base)},
            {'label': 'Tổng giá trị', 'value': fmt_price(total_value_base)},
            {'label': '', 'value': ''},
            {'label': '', 'value': ''},
        ]
    elif filter_type == 'delivered':
        stat_cards = [
            {'label': 'Tổng đơn hoàn thành', 'value': str(total_base)},
            {'label': 'Tổng doanh thu', 'value': fmt_price(total_revenue_base)},
            {'label': 'Tổng SP đã bán', 'value': str(items_sold_base) + ' SP'},
            {'label': '', 'value': ''},
        ]
    elif filter_type == 'cancelled':
        stat_cards = [
            {'label': 'Tổng đơn hủy', 'value': str(total_base)},
            {'label': 'Tổng giá trị bị hủy', 'value': fmt_price(total_value_base)},
            {'label': 'Tỷ lệ hủy tổng thể', 'value': str(cancel_rate) + '%'},
            {'label': '', 'value': ''},
        ]
    else:
        stat_cards = []

    # --- Áp dụng search + status sub-filter ---
    filtered_qs = base_qs
    if search:
        filtered_qs = filtered_qs.filter(order_code__icontains=search)
    if status_sub and is_time_filter:
        filtered_qs = filtered_qs.filter(status=status_sub)

    total = filtered_qs.count()
    total_pages = max(1, (total + per_page - 1) // per_page)
    if page > total_pages:
        page = total_pages

    start = (page - 1) * per_page
    paged_qs = list(filtered_qs[start: start + per_page])

    orders = []
    for idx, o in enumerate(paged_qs, start=start + 1):
        item_count = o.items.aggregate(t=Sum('quantity'))['t'] or 0
        orders.append({
            'id': o.id,
            'stt': idx,
            'order_code': o.order_code,
            'user_email': o.user.email if o.user else '-',
            'user_phone': (o.user.phone or '-') if o.user else '-',
            'total_amount': int(o.total_amount),
            'item_count': item_count,
            'status': o.status,
            'status_display': STATUS_DISPLAY.get(o.status, o.status),
            'payment_method': o.get_payment_method_display(),
            'created_at': o.created_at.strftime('%d/%m/%Y %H:%M'),
        })

    return JsonResponse({
        'success': True,
        'filter': filter_type,
        'is_time_filter': is_time_filter,
        'title': TITLE_MAP.get(filter_type, filter_type),
        'stat_cards': stat_cards,
        'total': total,
        'page': page,
        'total_pages': total_pages,
        'orders': orders,
    })



@login_required
@require_http_methods(["GET"])
def dashboard_product_detail(request):
    """API: Chi tiết sản phẩm theo filter cho dashboard stat boxes"""
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'message': 'Không có quyền!'}, status=403)

    from store.models import Product, OrderItem, Brand

    filter_type = request.GET.get('filter', 'all')  # all|active|outofstock|bestseller|lowstock
    page = max(1, int(request.GET.get('page', 1) or 1))
    search = request.GET.get('search', '').strip()
    per_page = 10

    TITLE_MAP = {
        'all': 'Tổng số sản phẩm',
        'active': 'Sản phẩm đang bán',
        'outofstock': 'Sản phẩm hết hàng',
        'bestseller': 'Top 10 sản phẩm bán chạy',
        'lowstock': 'Sản phẩm sắp hết hàng (tồn ≤ 5)',
    }

    def fmt_price(v): return f"{int(v):,}đ".replace(',', '.')

    if filter_type == 'bestseller':
        # Tổng SP bán ra theo order items
        from django.db.models import F
        top_qs = (
            OrderItem.objects
            .filter(product__isnull=False, order__status='delivered')
            .values('product_id', 'product__name', 'product__price', 'product__stock')
            .annotate(sold=Sum('quantity'), revenue=Sum(F('price') * F('quantity')))
            .order_by('-sold')[:50]
        )
        if search:
            top_qs = [r for r in top_qs if search.lower() in (r['product__name'] or '').lower()]

        total = len(top_qs)
        total_revenue_all = sum(r['revenue'] or 0 for r in top_qs)
        total_pages = max(1, (total + per_page - 1) // per_page)
        page = min(page, total_pages)
        paged = top_qs[(page - 1) * per_page: page * per_page]

        stat_cards = [
            {'label': 'Tổng sản phẩm có đơn', 'value': str(total)},
            {'label': 'Tổng doanh thu (tất cả)', 'value': fmt_price(total_revenue_all)},
            {'label': '', 'value': ''},
            {'label': '', 'value': ''},
        ]
        products = []
        for idx, r in enumerate(paged, start=(page - 1) * per_page + 1):
            pct = round(r['revenue'] / total_revenue_all * 100, 1) if total_revenue_all > 0 else 0
            products.append({
                'stt': idx,
                'name': r['product__name'] or '-',
                'price': int(r['product__price'] or 0),
                'stock': r['product__stock'],
                'sold': r['sold'],
                'revenue': int(r['revenue'] or 0),
                'pct_revenue': pct,
                'is_active': (r['product__stock'] or 0) > 0,
            })
        return JsonResponse({
            'success': True, 'filter': filter_type, 'title': TITLE_MAP[filter_type],
            'stat_cards': stat_cards, 'total': total, 'page': page,
            'total_pages': total_pages, 'products': products,
        })

    # Standard product queryset
    qs = Product.objects.select_related('brand').order_by('-created_at')
    if filter_type == 'active':
        qs = qs.filter(stock__gt=0)
    elif filter_type == 'outofstock':
        qs = qs.filter(stock=0)
    elif filter_type == 'lowstock':
        qs = qs.filter(stock__gt=0, stock__lte=5)

    if search:
        qs = qs.filter(name__icontains=search)

    total = qs.count()
    total_stock = qs.aggregate(s=Sum('stock'))['s'] or 0
    total_value = qs.aggregate(s=Sum('price'))['s'] or 0

    if filter_type == 'all':
        stat_cards = [
            {'label': 'Tổng số sản phẩm', 'value': str(total)},
            {'label': 'Tổng tồn kho', 'value': str(total_stock) + ' SP'},
            {'label': 'Hết hàng', 'value': str(qs.filter(stock=0).count())},
            {'label': 'Sắp hết (≤5)', 'value': str(qs.filter(stock__gt=0, stock__lte=5).count())},
        ]
    elif filter_type == 'active':
        stat_cards = [
            {'label': 'Đang bán', 'value': str(total)},
            {'label': 'Tổng tồn kho', 'value': str(total_stock) + ' SP'},
            {'label': '', 'value': ''},
            {'label': '', 'value': ''},
        ]
    elif filter_type == 'outofstock':
        stat_cards = [
            {'label': 'Hết hàng', 'value': str(total)},
            {'label': '', 'value': ''},
            {'label': '', 'value': ''},
            {'label': '', 'value': ''},
        ]
    elif filter_type == 'lowstock':
        stat_cards = [
            {'label': 'Sắp hết hàng', 'value': str(total)},
            {'label': 'Tổng tồn kho còn lại', 'value': str(total_stock) + ' SP'},
            {'label': '', 'value': ''},
            {'label': '', 'value': ''},
        ]
    else:
        stat_cards = []

    total_pages = max(1, (total + per_page - 1) // per_page)
    page = min(page, total_pages)
    paged_qs = list(qs[(page - 1) * per_page: page * per_page])

    products = []
    for idx, p in enumerate(paged_qs, start=(page - 1) * per_page + 1):
        products.append({
            'id': p.id,
            'stt': idx,
            'name': p.name,
            'brand': p.brand.name if p.brand else '-',
            'price': int(p.price),
            'original_price': int(p.original_price) if p.original_price else None,
            'stock': p.stock,
            'is_active': p.is_active,
        })

    return JsonResponse({
        'success': True, 'filter': filter_type, 'title': TITLE_MAP.get(filter_type, filter_type),
        'stat_cards': stat_cards, 'total': total, 'page': page,
        'total_pages': total_pages, 'products': products,
    })



def generate_slug(text):
    """Tạo slug từ text tiếng Việt"""
    # Remove accents
    import unicodedata
    text = unicodedata.normalize('NFD', text)
    text = text.encode('ascii', 'ignore').decode('ascii')
    # Convert to lowercase and replace spaces with hyphens
    text = re.sub(r'[^\w\s-]', '', text)
    text = re.sub(r'[-\s]+', '-', text)
    return text.lower().strip('-')



def _build_export_workbook(orders_qs, period_label, period_type):
    """
    Tạo workbook Excel 2 sheet: Danh sách đơn hàng + Báo cáo tổng hợp.
    period_type: 'month' hoặc 'year'
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from store.models import OrderItem, ProductVariant
    from django.db.models import Sum, Count

    STATUS_VN = {
        'awaiting_payment': 'Chờ thanh toán',
        'pending': 'Đã đặt hàng',
        'processing': 'Đang xử lý',
        'shipped': 'Đang giao',
        'delivered': 'Đã giao hàng',
        'cancelled': 'Đã hủy',
        'payment_expired': 'Hết hạn thanh toán',  # W4 FIX
    }
    PAYMENT_VN = {
        'cod': 'COD (Tiền mặt)',
        'vietqr': 'VietQR',
        'vnpay': 'VNPay',
        'momo': 'MoMo',
    }

    # Style helpers
    hdr_fill   = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
    hdr_font   = Font(color='FFFFFF', bold=True, size=10, name='Calibri')
    hdr_align  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin       = Side(style='thin', color='D1D5DB')
    border     = Border(left=thin, right=thin, top=thin, bottom=thin)
    money_fmt  = '#,##0'
    total_fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
    total_font = Font(bold=True, color='92400E', name='Calibri')
    section_fill = PatternFill(start_color='EFF6FF', end_color='EFF6FF', fill_type='solid')
    section_font = Font(bold=True, color='1D4ED8', name='Calibri', size=11)
    even_fill  = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')

    def hdr_cell(ws, row, col, value, width=None):
        c = ws.cell(row=row, column=col, value=value)
        c.fill = hdr_fill; c.font = hdr_font
        c.alignment = hdr_align; c.border = border
        return c

    def data_cell(ws, row, col, value, fmt=None, align='left', fill=None, bold=False):
        c = ws.cell(row=row, column=col, value=value)
        c.border = border
        c.alignment = Alignment(vertical='center', horizontal=align, wrap_text=False)
        if fmt: c.number_format = fmt
        if fill: c.fill = fill
        if bold: c.font = Font(bold=True, name='Calibri')
        return c

    # ── Lấy tất cả order items cùng một lúc ──────────────────────────────────
    orders_list = list(orders_qs.prefetch_related('items').select_related('user'))
    order_ids = [o.id for o in orders_list]
    all_items = list(
        OrderItem.objects.filter(order_id__in=order_ids)
        .select_related('order__user', 'product')
        .order_by('order_id')
    )

    # Batch-lookup variants để lấy SKU + giá gốc + % giảm
    product_ids = [it.product_id for it in all_items if it.product_id]
    variants = ProductVariant.objects.filter(
        detail__product_id__in=product_ids
    ).select_related('detail')
    variant_map = {}  # (product_id, color_name, storage) -> variant
    for v in variants:
        key = (v.detail.product_id, v.color_name.strip().lower(), v.storage.strip().lower())
        variant_map[key] = v

    # ══════════════════════════════════════════════════════
    # SHEET 1: DANH SÁCH ĐƠN HÀNG
    # ══════════════════════════════════════════════════════
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = 'Danh sách đơn hàng'
    ws1.freeze_panes = 'A2'

    S1_HEADERS = [
        ('STT', 5), ('Mã đơn hàng', 18), ('Tên khách hàng', 22),
        ('Email', 28), ('SĐT', 14), ('Tên sản phẩm', 36),
        ('SKU', 22), ('Phiên bản (RAM/ROM)', 18), ('Màu sắc', 14),
        ('Số lượng', 10), ('Giá gốc (đ)', 16), ('% Giảm', 8),
        ('Giá sau giảm (đ)', 16), ('Thành tiền (đ)', 16),
        ('Phương thức TT', 16), ('Trạng thái', 16),
        ('Ngày tạo', 18), ('Tháng', 7), ('Năm', 7),
    ]
    for col, (h, w) in enumerate(S1_HEADERS, 1):
        hdr_cell(ws1, 1, col, h)
        from openpyxl.utils import get_column_letter
        ws1.column_dimensions[get_column_letter(col)].width = w
    ws1.row_dimensions[1].height = 30

    row_num = 2
    stt = 0
    for order in orders_list:
        order_items = [it for it in all_items if it.order_id == order.id]
        if not order_items:
            # Đơn không có sản phẩm — vẫn ghi 1 dòng
            order_items = [None]
        for item in order_items:
            stt += 1
            is_even = (stt % 2 == 0)
            row_fill = even_fill if is_even else None

            if item:
                vkey = (
                    item.product_id,
                    item.color_name.strip().lower(),
                    item.storage.strip().lower(),
                ) if item.product_id else None
                variant = variant_map.get(vkey) if vkey else None
                sku         = variant.sku if variant else ''
                orig_price  = int(variant.original_price) if variant and variant.original_price else int(item.price)
                disc_pct    = variant.discount_percent if variant and variant.discount_percent else 0
                after_disc  = int(item.price)
                subtotal    = after_disc * item.quantity
                product_name = item.product_name
                storage      = item.storage or ''
                color        = item.color_name or ''
                qty          = item.quantity
            else:
                sku = orig_price = disc_pct = after_disc = subtotal = ''
                product_name = storage = color = ''
                qty = ''

            row_data = [
                (stt,           'center', None),
                (order.order_code, 'center', None),
                (f"{order.user.last_name} {order.user.first_name}".strip() if order.user else '', 'left', None),
                (order.user.email if order.user else '', 'left', None),
                (order.user.phone if order.user else '', 'center', None),
                (product_name,  'left', None),
                (sku,           'center', None),
                (storage,       'center', None),
                (color,         'center', None),
                (qty,           'center', None),
                (orig_price,    'right', money_fmt),
                (disc_pct,      'center', '0"%"'),
                (after_disc,    'right', money_fmt),
                (subtotal,      'right', money_fmt),
                (PAYMENT_VN.get(order.payment_method, order.payment_method), 'center', None),
                (STATUS_VN.get(order.status, order.status), 'center', None),
                (order.created_at.strftime('%d/%m/%Y %H:%M'), 'center', None),
                (order.created_at.month, 'center', None),
                (order.created_at.year,  'center', None),
            ]
            for col, (val, align, fmt) in enumerate(row_data, 1):
                data_cell(ws1, row_num, col, val, fmt=fmt, align=align, fill=row_fill)
            row_num += 1

    # Dòng tổng
    total_revenue_all = sum(int(o.total_amount) for o in orders_list if o.status == 'delivered')
    tr = row_num
    for c in range(1, 20):
        cell = ws1.cell(row=tr, column=c)
        cell.fill = total_fill; cell.border = border
    ws1.cell(row=tr, column=1, value='TỔNG DOANH THU (đơn đã giao)').font = total_font
    ws1.cell(row=tr, column=1).fill = total_fill
    ws1.merge_cells(start_row=tr, start_column=1, end_row=tr, end_column=13)
    rev = ws1.cell(row=tr, column=14, value=total_revenue_all)
    rev.font = total_font; rev.fill = total_fill
    rev.number_format = money_fmt
    rev.alignment = Alignment(horizontal='right', vertical='center')

    # ══════════════════════════════════════════════════════
    # SHEET 2: BÁO CÁO TỔNG HỢP
    # ══════════════════════════════════════════════════════
    ws2 = wb.create_sheet('Báo cáo tổng hợp')
    ws2.column_dimensions['A'].width = 32
    ws2.column_dimensions['B'].width = 22
    ws2.column_dimensions['C'].width = 22
    ws2.column_dimensions['D'].width = 22

    def section_title(ws, row, title):
        c = ws.cell(row=row, column=1, value=title)
        c.font = section_font; c.fill = section_fill
        c.alignment = Alignment(vertical='center')
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        ws.row_dimensions[row].height = 22
        return row + 1

    def kv_row(ws, row, label, value, fmt=None, highlight=False):
        lc = ws.cell(row=row, column=1, value=label)
        vc = ws.cell(row=row, column=2, value=value)
        lc.font = Font(name='Calibri')
        vc.font = Font(bold=highlight, name='Calibri', color='DC2626' if highlight else '000000')
        vc.alignment = Alignment(horizontal='right', vertical='center')
        if fmt: vc.number_format = fmt
        lc.border = border; vc.border = border
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        if highlight:
            for c in range(1, 5):
                ws.cell(row=row, column=c).fill = total_fill
        return row + 1

    r = 1
    # --- Tiêu đề báo cáo ---
    title_cell = ws2.cell(row=r, column=1, value=f'BÁO CÁO DOANH THU — {period_label.upper()}')
    title_cell.font = Font(bold=True, size=14, color='1E293B', name='Calibri')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    ws2.row_dimensions[r].height = 28
    r += 2

    # --- 1. Tổng quan ---
    r = section_title(ws2, r, '1. TỔNG QUAN')
    total_orders  = len(orders_list)
    total_qty     = sum(it.quantity for it in all_items)
    total_revenue = sum(int(o.total_amount) for o in orders_list if o.status == 'delivered')
    delivered_cnt = sum(1 for o in orders_list if o.status == 'delivered')
    # W2 FIX: Excel báo cáo cancelled_cnt cần bao gồm cả payment_expired (không thanh toán)
    cancelled_cnt = sum(1 for o in orders_list if o.status in ('cancelled', 'payment_expired'))

    r = kv_row(ws2, r, 'Tổng số đơn hàng',   total_orders)
    r = kv_row(ws2, r, 'Tổng sản phẩm đã bán', total_qty)
    r = kv_row(ws2, r, 'Tổng số đơn đã giao', delivered_cnt)
    r = kv_row(ws2, r, 'Tổng số đơn đã hủy',  cancelled_cnt)
    r = kv_row(ws2, r, 'Tổng doanh thu (đơn đã giao)', total_revenue, fmt=money_fmt, highlight=True)
    r += 1

    # --- 2. Doanh thu theo ngày hoặc tháng ---
    if period_type == 'month':
        r = section_title(ws2, r, '2. DOANH THU THEO NGÀY')
        hdr_cell(ws2, r, 1, 'Ngày'); hdr_cell(ws2, r, 2, 'Số đơn')
        hdr_cell(ws2, r, 3, 'Doanh thu (đ)'); hdr_cell(ws2, r, 4, 'Đơn đã giao')
        r += 1
        by_day = {}
        for o in orders_list:
            d = o.created_at.day
            if d not in by_day:
                by_day[d] = {'count': 0, 'revenue': 0, 'delivered': 0}
            by_day[d]['count'] += 1
            if o.status == 'delivered':
                by_day[d]['revenue'] += int(o.total_amount)
                by_day[d]['delivered'] += 1
        for day in sorted(by_day):
            dd = by_day[day]
            data_cell(ws2, r, 1, f'Ngày {day:02d}', align='center')
            data_cell(ws2, r, 2, dd['count'], align='center')
            data_cell(ws2, r, 3, dd['revenue'], fmt=money_fmt, align='right')
            data_cell(ws2, r, 4, dd['delivered'], align='center')
            r += 1
    else:
        r = section_title(ws2, r, '2. DOANH THU THEO THÁNG')
        hdr_cell(ws2, r, 1, 'Tháng'); hdr_cell(ws2, r, 2, 'Số đơn')
        hdr_cell(ws2, r, 3, 'Doanh thu (đ)'); hdr_cell(ws2, r, 4, 'Đơn đã giao')
        r += 1
        by_month = {m: {'count': 0, 'revenue': 0, 'delivered': 0} for m in range(1, 13)}
        for o in orders_list:
            m = o.created_at.month
            by_month[m]['count'] += 1
            if o.status == 'delivered':
                by_month[m]['revenue'] += int(o.total_amount)
                by_month[m]['delivered'] += 1
        for m in range(1, 13):
            dd = by_month[m]
            data_cell(ws2, r, 1, f'Tháng {m:02d}', align='center')
            data_cell(ws2, r, 2, dd['count'], align='center')
            data_cell(ws2, r, 3, dd['revenue'], fmt=money_fmt, align='right')
            data_cell(ws2, r, 4, dd['delivered'], align='center')
            r += 1
    r += 1

    # --- 3. Top sản phẩm bán chạy ---
    r = section_title(ws2, r, '3. TOP SẢN PHẨM BÁN CHẠY')
    hdr_cell(ws2, r, 1, 'Tên sản phẩm')
    hdr_cell(ws2, r, 2, 'Tổng SL bán')
    hdr_cell(ws2, r, 3, 'Tổng doanh thu (đ)')
    hdr_cell(ws2, r, 4, 'Số đơn hàng')
    r += 1
    product_stats = {}
    for it in all_items:
    # W3 FIX: loại trừ cả 'cancelled' và 'payment_expired' khỏi top products
        if it.order.status in ('cancelled', 'payment_expired'):
            continue
        key = it.product_name
        if key not in product_stats:
            product_stats[key] = {'qty': 0, 'revenue': 0, 'orders': set()}
        product_stats[key]['qty'] += it.quantity
        product_stats[key]['revenue'] += int(it.price) * it.quantity
        product_stats[key]['orders'].add(it.order_id)
    top_products = sorted(product_stats.items(), key=lambda x: x[1]['qty'], reverse=True)[:20]
    for i, (name, stat) in enumerate(top_products):
        fill = even_fill if i % 2 == 0 else None
        data_cell(ws2, r, 1, name, align='left', fill=fill)
        data_cell(ws2, r, 2, stat['qty'], align='center', fill=fill)
        data_cell(ws2, r, 3, stat['revenue'], fmt=money_fmt, align='right', fill=fill)
        data_cell(ws2, r, 4, len(stat['orders']), align='center', fill=fill)
        r += 1

    return wb


@login_required
@require_http_methods(["GET"])
def export_revenue_month(request):
    """Xuất báo cáo tháng ra file .xlsx (2 sheet)"""
    if not request.user.is_superuser:
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden('Không có quyền!')

    from django.http import HttpResponse
    from store.models import Order

    now = timezone.now()
    try:
        month = int(request.GET.get('month', now.month))
        year  = int(request.GET.get('year',  now.year))
        if not (1 <= month <= 12): month = now.month
        if not (2000 <= year <= 2100): year = now.year
    except (ValueError, TypeError):
        month, year = now.month, now.year

    qs = Order.objects.filter(
        created_at__year=year, created_at__month=month
    ).order_by('created_at')

    period_label = f'Tháng {month:02d}/{year}'
    wb = _build_export_workbook(qs, period_label, 'month')

    filename = f'bao-cao-thang-{month:02d}-{year}.xlsx'
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response



@login_required
@require_POST
def dashboard_save_cost_price(request):
    """Lưu giá vốn nhập hàng cho từng sản phẩm (AJAX)."""
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'error': 'Không có quyền!'}, status=403)

    from store.models import Product
    try:
        data = json.loads(request.body)
        items = data.get('items', [])
        if not isinstance(items, list):
            return JsonResponse({'success': False, 'error': 'Dữ liệu không hợp lệ'}, status=400)

        updated = 0
        for item in items:
            pid = item.get('id')
            raw_cost = item.get('cost_price')
            if pid is None:
                continue
            try:
                pid = int(pid)
                cost = None if (raw_cost is None or str(raw_cost).strip() == '') else int(str(raw_cost).replace(',', '').replace('.', ''))
                Product.objects.filter(id=pid).update(cost_price=cost)
                updated += 1
            except (ValueError, TypeError):
                continue

        return JsonResponse({'success': True, 'updated': updated})
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'JSON không hợp lệ'}, status=400)


@login_required
@require_http_methods(["GET"])
def export_revenue_year(request):
    """Xuất báo cáo năm ra file .xlsx (2 sheet)"""
    if not request.user.is_superuser:
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden('Không có quyền!')

    from django.http import HttpResponse
    from store.models import Order

    now = timezone.now()
    try:
        year = int(request.GET.get('year', now.year))
        if not (2000 <= year <= 2100): year = now.year
    except (ValueError, TypeError):
        year = now.year

    qs = Order.objects.filter(created_at__year=year).order_by('created_at')

    period_label = f'Năm {year}'
    wb = _build_export_workbook(qs, period_label, 'year')

    filename = f'bao-cao-nam-{year}.xlsx'
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response



@login_required
def brand_list(request):
    """Danh sách hãng sản phẩm"""
    if not request.user.is_superuser:
        messages.error(request, 'Bạn không có quyền truy cập!')
        return redirect('store:profile')
    
    from store.models import Brand
    brands = Brand.objects.all().order_by('name')
    
    # Search
    search = request.GET.get('search', '')
    if search:
        brands = brands.filter(name__icontains=search)
    
    # Pagination (8 hãng / trang)
    brand_page = request.GET.get('brand_page', 1)
    brand_paginator = Paginator(brands, 8)
    try:
        brands_paginated = brand_paginator.page(brand_page)
    except PageNotAnInteger:
        brands_paginated = brand_paginator.page(1)
    except EmptyPage:
        brands_paginated = brand_paginator.page(brand_paginator.num_pages)
    
    context = {
        'brands_paginated': brands_paginated,
        'search': search,
    }
    return render(request, 'store/admin/brand_list.html', context)



@login_required
@require_http_methods(["POST"])
def brand_add(request):
    """Thêm hãng mới"""
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'message': 'Không có quyền!'}, status=403)
    
    from store.models import Brand
    name = request.POST.get('name', '').strip()
    description = request.POST.get('description', '').strip()
    
    if not name:
        return JsonResponse({'success': False, 'message': 'Tên hãng không được để trống!'}, status=400)
    
    # Check exists
    if Brand.objects.filter(name__iexact=name).exists():
        return JsonResponse({'success': False, 'message': 'Hãng đã tồn tại!'}, status=400)
    
    slug = generate_slug(name)
    # Ensure unique slug
    base_slug = slug
    counter = 1
    while Brand.objects.filter(slug=slug).exists():
        slug = f"{base_slug}-{counter}"
        counter += 1
    
    brand = Brand.objects.create(
        name=name,
        slug=slug,
        description=description,
        is_active=True
    )
    
    return JsonResponse({
        'success': True, 
        'message': f'Đã thêm hãng "{name}"!',
        'brand': {'id': brand.id, 'name': brand.name, 'slug': brand.slug}
    })



@login_required
@require_http_methods(["POST"])
def brand_edit(request):
    """Sửa hãng"""
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'message': 'Không có quyền!'}, status=403)
    
    from store.models import Brand
    brand_id = request.POST.get('brand_id')
    name = request.POST.get('name', '').strip()
    description = request.POST.get('description', '').strip()
    
    if not name or not brand_id:
        return JsonResponse({'success': False, 'message': 'Thiếu thông tin!'}, status=400)
    
    try:
        brand = Brand.objects.get(id=brand_id)
    except Brand.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Hãng không tồn tại!'}, status=404)
    
    # Check duplicate name (excluding current brand)
    if Brand.objects.filter(name__iexact=name).exclude(id=brand_id).exists():
        return JsonResponse({'success': False, 'message': 'Tên hãng đã tồn tại!'}, status=400)
    
    brand.name = name
    brand.description = description
    brand.slug = generate_slug(name)
    brand.save()
    
    return JsonResponse({'success': True, 'message': f'Đã cập nhật hãng "{name}"!'})



@login_required
@require_http_methods(["POST"])
def brand_delete(request):
    """Xóa hãng"""
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'message': 'Không có quyền!'}, status=403)
    
    from store.models import Brand
    brand_id = request.POST.get('brand_id')
    
    if not brand_id:
        return JsonResponse({'success': False, 'message': 'Thiếu thông tin!'}, status=400)
    
    try:
        brand = Brand.objects.get(id=brand_id)
    except Brand.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Hãng không tồn tại!'}, status=404)
    
    # Check if brand has products
    if brand.products.exists():
        return JsonResponse({'success': False, 'message': 'Không thể xóa! Hãng đang có sản phẩm.'}, status=400)
    
    brand_name = brand.name
    brand.delete()
    
    return JsonResponse({'success': True, 'message': f'Đã xóa hãng "{brand_name}"!'})


@login_required
def user_detail_json(request):
    """Trả về JSON thông tin chi tiết người dùng + sổ địa chỉ"""
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'message': 'Không có quyền!'}, status=403)
    user_id = request.GET.get('user_id')
    if not user_id:
        return JsonResponse({'success': False, 'message': 'Thiếu user_id!'}, status=400)
    from store.models import CustomUser, Address
    try:
        u = CustomUser.objects.get(id=user_id)
    except CustomUser.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Người dùng không tồn tại!'}, status=404)
    addresses = Address.objects.filter(user=u)
    addr_list = [{
        'full_name': a.full_name,
        'phone': a.phone,
        'detail': a.detail,
        'ward_name': a.ward_name,
        'district_name': a.district_name,
        'province_name': a.province_name,
        'is_default': a.is_default,
    } for a in addresses]
    return JsonResponse({
        'success': True,
        'user': {
            'id': u.id,
            'email': u.email,
            'last_name': u.last_name or '',
            'first_name': u.first_name or '',
            'phone': u.phone or '',
            'verified_student_email': u.verified_student_email or '',
            'verified_teacher_email': u.verified_teacher_email or '',
        },
        'addresses': addr_list,
    })


@login_required
@require_http_methods(["POST"])
@csrf_exempt
def user_add(request):
    """Thêm người dùng mới"""
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'message': 'Không có quyền!'}, status=403)

    from store.models import CustomUser
    email = request.POST.get('email', '').strip().lower()
    phone = request.POST.get('phone', '').strip()
    last_name = request.POST.get('last_name', '').strip()
    first_name = request.POST.get('first_name', '').strip()
    password = request.POST.get('password', '').strip()

    if not email or not password:
        return JsonResponse({'success': False, 'message': 'Email và mật khẩu không được để trống!'}, status=400)

    if CustomUser.objects.filter(email=email).exists():
        return JsonResponse({'success': False, 'message': 'Email này đã tồn tại!'}, status=400)

    user = CustomUser.objects.create_user(
        username=email,
        email=email,
        password=password,
        last_name=last_name,
        first_name=first_name,
    )
    if phone:
        user.phone = phone
        user.save()

    return JsonResponse({'success': True, 'message': f'Tạo tài khoản "{email}" thành công!'})



@login_required
@require_http_methods(["POST"])
@csrf_exempt
def user_edit(request):
    """Sửa thông tin người dùng"""
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'message': 'Không có quyền!'}, status=403)
    
    from store.models import CustomUser
    user_id = request.POST.get('user_id')
    last_name = request.POST.get('last_name', '').strip()
    first_name = request.POST.get('first_name', '').strip()
    phone = request.POST.get('phone', '').strip()
    
    if not user_id:
        return JsonResponse({'success': False, 'message': 'Thiếu thông tin!'}, status=400)
    
    try:
        user = CustomUser.objects.get(id=user_id)
    except CustomUser.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Người dùng không tồn tại!'}, status=404)
    
    user.last_name = last_name
    user.first_name = first_name
    user.phone = phone if phone else None
    user.save()
    
    return JsonResponse({'success': True, 'message': 'Cập nhật thông tin thành công!'})


@login_required
@require_http_methods(["POST"])
@csrf_exempt
def user_delete(request):
    """Xóa người dùng"""
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'message': 'Không có quyền!'}, status=403)
    
    from store.models import CustomUser
    user_id = request.POST.get('user_id')
    
    if not user_id:
        return JsonResponse({'success': False, 'message': 'Thiếu thông tin!'}, status=400)
    
    try:
        user = CustomUser.objects.get(id=user_id)
    except CustomUser.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Người dùng không tồn tại!'}, status=404)
    
    # Không cho xóa admin
    if user.is_superuser:
        return JsonResponse({'success': False, 'message': 'Không thể xóa tài khoản admin!'}, status=400)
    
    user_email = user.email
    user.delete()
    
    return JsonResponse({'success': True, 'message': f'Đã xóa người dùng "{user_email}"!'})


@login_required
@require_http_methods(["POST"])
@csrf_exempt
def product_edit(request):
    """Sửa sản phẩm (Product)"""
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'message': 'Không có quyền!'}, status=403)
    
    from store.models import Product, Brand
    
    product_id = request.POST.get('product_id')
    brand_id = request.POST.get('brand')
    name = request.POST.get('name', '').strip()
    
    if not product_id or not name:
        return JsonResponse({'success': False, 'message': 'Thiếu thông tin cần thiết!'}, status=400)
    
    try:
        product = Product.objects.get(id=product_id)
    except Product.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Sản phẩm không tồn tại!'}, status=404)
    
    try:
        brand = Brand.objects.get(id=brand_id) if brand_id else None
    except Brand.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Hãng không tồn tại!'}, status=404)
    
    # Handle image upload - if new image uploaded, delete old and save new
    new_image = product.image
    new_image_file = request.FILES.get('image')
    
    if new_image_file:
        # Delete old local image
        if product.image:
            try:
                import os
                if os.path.isfile(product.image.path):
                    os.remove(product.image.path)
                product.image.delete(save=False)
            except Exception as e:
                print(f"Error deleting old image: {e}")
        new_image = new_image_file
    
    product.brand = brand
    product.name = name
    product.image = new_image
    product.save()
    
    return JsonResponse({'success': True, 'message': f'Đã cập nhật sản phẩm "{name}"!'})



@login_required
@require_http_methods(["POST"])
@csrf_exempt
def product_delete(request):
    """Xóa sản phẩm (Product)"""
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'message': 'Không có quyền!'}, status=403)
    
    from store.models import Product
    
    product_id = request.POST.get('product_id')
    
    if not product_id:
        return JsonResponse({'success': False, 'message': 'Thiếu thông tin!'}, status=400)
    
    try:
        product = Product.objects.get(id=product_id)
    except Product.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Sản phẩm không tồn tại!'}, status=404)
    
    product_name = product.name
    product.delete()
    
    return JsonResponse({'success': True, 'message': f'Đã xóa sản phẩm "{product_name}"!'})

@login_required
@require_http_methods(["POST"])
@csrf_exempt
def product_add(request):
    """Thêm sản phẩm mới"""
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'message': 'Không có quyền!'}, status=403)
    
    from store.models import Product, Brand
    import os
    
    brand_id = request.POST.get('brand')
    name = request.POST.get('name', '').strip()
    
    # Validation
    if not brand_id or not name:
        return JsonResponse({'success': False, 'message': 'Thiếu thông tin cần thiết!'}, status=400)
    
    try:
        brand = Brand.objects.get(id=brand_id)
    except Brand.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Hãng không tồn tại!'}, status=404)
    
    # Handle image upload
    image = request.FILES.get('image')
    
    # Generate slug from name
    from django.utils.text import slugify
    slug = slugify(name)
    
    # Ensure unique slug
    base_slug = slug
    counter = 1
    while Product.objects.filter(slug=slug).exists():
        slug = f"{base_slug}-{counter}"
        counter += 1
    
    product = Product.objects.create(
        brand=brand,
        name=name,
        slug=slug,
        description='',
        image=image,
        is_active=True
    )
    
    return JsonResponse({'success': True, 'message': f'Đã thêm sản phẩm "{name}"!'})

@login_required
@require_http_methods(["POST"])
@csrf_exempt
def product_detail_save(request):
    """Lưu chi tiết sản phẩm (tạo/cập nhật Product và ProductDetail)"""
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'message': 'Không có quyền!'}, status=403)
    
    from store.models import Product, ProductDetail, ProductVariant, FolderColorImage
    
    product_id = request.POST.get('product_id')
    original_price = request.POST.get('original_price', '0').strip()
    discount_percent = request.POST.get('discount_percent', '0').strip()
    stock = request.POST.get('stock', '0').strip()
    sku = request.POST.get('sku', '').strip()
    save_sku_only = request.POST.get('save_sku_only', 'false').strip().lower() == 'true'
    delete_sku = request.POST.get('delete_sku', 'false').strip().lower() == 'true'
    
    if not product_id:
        return JsonResponse({'success': False, 'message': 'Thiếu thông tin sản phẩm!'}, status=400)
    
    try:
        product = Product.objects.get(id=product_id)
    except Product.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Sản phẩm không tồn tại!'}, status=404)
    
    # Get or create ProductDetail
    detail = ProductDetail.objects.filter(product=product).first()
    
    # Handle delete SKU
    if delete_sku:
        sku_to_delete = request.POST.get('sku', '').strip()
        if not detail or not detail.sku:
            return JsonResponse({'success': False, 'message': 'Không có SKU để xóa!'}, status=400)
        
        existing_skus = detail.sku.split(',') if detail.sku else []
        if sku_to_delete in existing_skus:
            existing_skus.remove(sku_to_delete)
            detail.sku = ','.join(existing_skus)
            detail.save(update_fields=['sku'])
        
        return JsonResponse({
            'success': True, 
            'message': f'Đã xóa SKU: {sku_to_delete}!',
            'sku': detail.sku
        })
    
    # If save_sku_only, only save SKU
    if save_sku_only:
        if not sku:
            return JsonResponse({'success': False, 'message': 'Vui lòng nhập SKU!'}, status=400)
        
        if not detail:
            # Check for duplicate in case detail was just created but we have a new request
            # Create new ProductDetail
            detail = ProductDetail.objects.create(product=product, sku=sku)
        else:
            # Append SKU if not already in list (comma separated)
            existing_skus = detail.sku.split(',') if detail.sku else []
            if sku in existing_skus:
                return JsonResponse({
                    'success': False, 
                    'message': f'SKU "{sku}" đã tồn tại!',
                }, status=400)
            existing_skus.append(sku)
            detail.sku = ','.join(existing_skus)
            detail.save(update_fields=['sku'])
        
        return JsonResponse({
            'success': True, 
            'message': f'Đã lưu SKU: {sku}!',
            'sku': detail.sku
        })
    
    # Save all fields
    try:
        original_price = int(original_price) if original_price else 0
        discount_percent = int(discount_percent) if discount_percent else 0
        stock = int(stock) if stock else 0
    except ValueError:
        return JsonResponse({'success': False, 'message': 'Giá phải là số!'}, status=400)
    
    # Calculate discounted price
    discounted_price = original_price - (original_price * discount_percent / 100)
    if discounted_price >= 5000:
        discounted_price = round(discounted_price / 5000) * 5000
    
    # Save to Product model
    product.original_price = original_price
    product.discount_percent = discount_percent
    product.price = discounted_price
    product.stock = stock
    product.save(update_fields=['original_price', 'discount_percent', 'price', 'stock'])
    
    # Save to ProductDetail model (for template display)
    if not detail:
        detail = ProductDetail.objects.create(
            product=product,
            original_price=original_price,
            discount_percent=discount_percent,
            sku=sku
        )
    else:
        detail.original_price = original_price
        detail.discount_percent = discount_percent
        if sku:
            # Append SKU if not already in list
            existing_skus = detail.sku.split(',') if detail.sku else []
            if sku not in existing_skus:
                existing_skus.append(sku)
                detail.sku = ','.join(existing_skus)
        detail.save(update_fields=['original_price', 'discount_percent', 'sku'])
    
    return JsonResponse({'success': True, 'message': f'Đã lưu thông tin sản phẩm!', 'detail_id': detail.id})



@login_required
@require_http_methods(["POST"])
@csrf_exempt
def product_variant_save(request):
    """Lưu biến thể sản phẩm (màu + dung lượng)"""
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'message': 'Không có quyền!'}, status=403)
    
    from store.models import Product, ProductDetail, ProductVariant, FolderColorImage
    
    variant_id = request.POST.get('variant_id')
    detail_id = request.POST.get('detail_id')
    product_id = request.POST.get('product_id')
    color_name = request.POST.get('color_name', '').strip()
    color_hex = request.POST.get('color_hex', '').strip()
    storage = request.POST.get('storage', '').strip()
    original_price = request.POST.get('original_price', request.POST.get('base_price', '0')).strip()
    discount_percent = request.POST.get('discount_percent', '0').strip()
    price = request.POST.get('price', '0').strip()
    variant_sku = request.POST.get('sku', '').strip()
    stock_quantity = request.POST.get('stock_quantity', '0').strip()
    
    if not color_name or not storage:
        return JsonResponse({'success': False, 'message': 'Thiếu thông tin biến thể!'}, status=400)
    
    # Xác định ProductDetail (detail) từ detail_id hoặc product_id
    detail = None
    if detail_id:
        try:
            detail = ProductDetail.objects.get(id=detail_id)
        except ProductDetail.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Chi tiết sản phẩm không tồn tại!'}, status=404)
    else:
        if not product_id:
            return JsonResponse({'success': False, 'message': 'Thiếu thông tin sản phẩm!'}, status=400)
        try:
            product = Product.objects.get(id=product_id)
        except Product.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Sản phẩm không tồn tại!'}, status=404)
        detail, _ = ProductDetail.objects.get_or_create(product=product)
    
    try:
        price = int(price) if price else 0
        stock_quantity = int(stock_quantity) if stock_quantity else 0
        original_price = int(original_price) if original_price else 0
        discount_percent = min(100, max(0, int(discount_percent) if discount_percent else 0))
    except ValueError:
        return JsonResponse({'success': False, 'message': 'Giá và số lượng phải là số!'}, status=400)
    
    if variant_id:
        # Update existing variant
        try:
            variant = ProductVariant.objects.get(id=variant_id, detail=detail)
            variant.color_name = color_name
            variant.color_hex = color_hex
            variant.storage = storage
            variant.original_price = original_price
            variant.discount_percent = discount_percent
            variant.price = price
            variant.sku = variant_sku
            variant.stock_quantity = stock_quantity
            variant.save()
            return JsonResponse({'success': True, 'message': f'Đã cập nhật biến thể "{color_name} - {storage}"!', 'variant_id': variant.id})
        except ProductVariant.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Biến thể không tồn tại!'}, status=404)
    else:
        # Create new variant
        variant, created = ProductVariant.objects.get_or_create(
            detail=detail,
            color_name=color_name,
            storage=storage,
            defaults={
                'color_hex': color_hex,
                'original_price': original_price,
                'discount_percent': discount_percent,
                'price': price,
                'sku': variant_sku,
                'stock_quantity': stock_quantity
            }
        )
        if not created:
            return JsonResponse({'success': False, 'message': 'Biến thể đã tồn tại!'}, status=400)
        return JsonResponse({'success': True, 'message': f'Đã thêm biến thể "{color_name} - {storage}"!', 'variant_id': variant.id})



@login_required
@require_http_methods(["POST"])
@csrf_exempt
def product_variant_delete(request):
    """Xóa biến thể sản phẩm"""
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'message': 'Không có quyền!'}, status=403)
    
    from store.models import ProductVariant
    
    variant_id = request.POST.get('variant_id')
    
    if not variant_id:
        return JsonResponse({'success': False, 'message': 'Thiếu thông tin!'}, status=400)
    
    try:
        variant = ProductVariant.objects.get(id=variant_id)
        variant_name = f"{variant.color_name} - {variant.storage}"
        variant.delete()
        return JsonResponse({'success': True, 'message': f'Đã xóa biến thể "{variant_name}"!'})
    except ProductVariant.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Biến thể không tồn tại!'}, status=404)



@login_required
@require_http_methods(["POST"])
@csrf_exempt
def product_image_upload(request):
    """Upload ảnh sản phẩm"""
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'message': 'Không có quyền!'}, status=403)
    
    from store.models import ProductDetail, ProductVariant, ProductImage
    import os
    from django.utils.text import slugify
    from datetime import datetime
    
    detail_id = request.POST.get('detail_id')
    variant_id = request.POST.get('variant_id')
    image_type = request.POST.get('image_type', 'cover')
    
    images = request.FILES.getlist('images')
    
    if not images:
        return JsonResponse({'success': False, 'message': 'Chưa chọn ảnh!'}, status=400)
    
    # Get product detail
    detail = None
    if detail_id:
        try:
            detail = ProductDetail.objects.get(id=detail_id)
        except ProductDetail.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Chi tiết sản phẩm không tồn tại!'}, status=404)
    
    # Get variant if provided
    variant = None
    if variant_id:
        try:
            variant = ProductVariant.objects.get(id=variant_id)
        except ProductVariant.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Biến thể không tồn tại!'}, status=404)
    
    # Create upload path
    product_slug = slugify(detail.product.name) if detail else slugify(variant.detail.product.name)
    year = datetime.now().year
    month = datetime.now().strftime('%m')
    
    uploaded_count = 0
    for img in images:
        # Get next order number
        max_order = ProductImage.objects.filter(detail=detail, variant=variant, image_type=image_type).aggregate(Max('order'))['order__max'] or 0
        
        # Create image record
        image = ProductImage.objects.create(
            detail=detail,
            variant=variant,
            image_type=image_type,
            image=img,
            order=max_order + 1
        )
        uploaded_count += 1
    
    return JsonResponse({'success': True, 'message': f'Đã upload {uploaded_count} ảnh!'})



@login_required
@require_http_methods(["POST"])
@csrf_exempt
def product_image_delete(request):
    """Xóa ảnh sản phẩm"""
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'message': 'Không có quyền!'}, status=403)
    
    from store.models import ProductImage
    import os
    
    image_id = request.POST.get('image_id')
    
    if not image_id:
        return JsonResponse({'success': False, 'message': 'Thiếu thông tin!'}, status=400)
    
    try:
        img = ProductImage.objects.get(id=image_id)
        # Delete file
        if img.image:
            try:
                if os.path.isfile(img.image.path):
                    os.remove(img.image.path)
                img.image.delete(save=True)
            except Exception as e:
                print(f"Error deleting image file: {e}")
        img.delete()
        return JsonResponse({'success': True, 'message': 'Đã xóa ảnh!'})
    except ProductImage.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Ảnh không tồn tại!'}, status=404)



@login_required
@require_http_methods(["GET"])
def image_folder_list(request):
    """Danh sách màu ảnh theo thư mục (group theo thư mục + màu + SKU)"""
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'message': 'Không có quyền!'}, status=403)

    from store.models import ImageFolder, FolderColorImage

    search = request.GET.get('search', '').strip()
    brand_id = request.GET.get('brand_id', '').strip()

    try:
        images_qs = FolderColorImage.objects.select_related('folder', 'folder__brand', 'folder__product', 'brand')
        if search:
            images_qs = images_qs.filter(folder__name__icontains=search)

        # Group theo (folder, color_name, sku)
        rows_map = {}
        for img in images_qs.order_by('folder__name', 'color_name', 'sku', 'order'):
            key = (img.folder_id, img.color_name, img.sku)
            if key not in rows_map:
                rows_map[key] = {
                    'id': img.id,
                    'folder_id': img.folder_id,
                    'folder_name': img.folder.name,
                    'color_name': img.color_name,
                    'sku': img.sku,
                    'brand_id': img.brand.id if img.brand else None,
                    'brand_name': img.brand.name if img.brand else None,
                    'folder_brand_id': img.folder.brand.id if img.folder.brand else None,
                    'folder_product_id': img.folder.product.id if img.folder.product else None,
                }

        rows = list(rows_map.values())

        # Danh sách tất cả thư mục (cho dropdown)
        folders_qs = ImageFolder.objects.select_related('brand', 'product').all().order_by('-created_at')
        
        # Lọc theo brand_id nếu có
        if brand_id:
            folders_qs = folders_qs.filter(brand_id=brand_id)
        
        folders = [{
            'id': f.id, 
            'name': f.name,
            'brand_id': f.brand.id if f.brand else None,
            'brand_name': f.brand.name if f.brand else None,
            'product_id': f.product.id if f.product else None,
            'product_name': f.product.name if f.product else None
        } for f in folders_qs]

        return JsonResponse({'success': True, 'rows': rows, 'folders': folders})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)



@login_required
@require_http_methods(["POST"])
@csrf_exempt
def image_folder_create(request):
    """Tạo thư mục ảnh mới"""
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'message': 'Không có quyền!'}, status=403)

    from store.models import ImageFolder, Brand, Product

    name = request.POST.get('name', '').strip()
    brand_id = request.POST.get('brand_id', '').strip()
    product_id = request.POST.get('product_id', '').strip()
    
    if not name:
        return JsonResponse({'success': False, 'message': 'Vui lòng nhập tên thư mục!'}, status=400)
    
    if not brand_id:
        return JsonResponse({'success': False, 'message': 'Vui lòng chọn hãng!'}, status=400)
    
    if not product_id:
        return JsonResponse({'success': False, 'message': 'Vui lòng chọn sản phẩm!'}, status=400)

    try:
        brand = Brand.objects.get(id=brand_id)
    except Brand.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Hãng không tồn tại!'}, status=404)
    
    try:
        product = Product.objects.get(id=product_id)
    except Product.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Sản phẩm không tồn tại!'}, status=404)

    try:
        folder, created = ImageFolder.objects.get_or_create(
            name=name,
            brand=brand,
            product=product
        )
        now = timezone.now()
        year = now.year
        month = now.strftime('%m')
        dir_path = os.path.join(settings.MEDIA_ROOT, 'products', str(year), month, folder.slug)
        try:
            os.makedirs(dir_path, exist_ok=True)
        except OSError as e:
            return JsonResponse({'success': False, 'message': f'Không tạo được thư mục trên đĩa: {e}'}, status=500)
        if created:
            message = f'Đã tạo thư mục \"{folder.name}\" (đường dẫn: media/products/{year}/{month}/{folder.slug}/).'
        else:
            message = f'Thư mục \"{folder.name}\" đã tồn tại.'
        return JsonResponse({
            'success': True,
            'message': message,
            'folder': {
                'id': folder.id, 
                'name': folder.name, 
                'slug': folder.slug,
                'brand_id': folder.brand.id if folder.brand else None,
                'brand_name': folder.brand.name if folder.brand else None,
                'product_id': folder.product.id if folder.product else None,
                'product_name': folder.product.name if folder.product else None
            }
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)



@login_required
@require_http_methods(["POST"])
@csrf_exempt
def image_folder_rename(request):
    """Đổi tên thư mục ảnh"""
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'message': 'Không có quyền!'}, status=403)

    from store.models import ImageFolder

    folder_id = request.POST.get('folder_id', '').strip()
    new_name = request.POST.get('name', '').strip()
    if not folder_id or not new_name:
        return JsonResponse({'success': False, 'message': 'Thiếu folder_id hoặc tên mới!'}, status=400)

    try:
        folder = ImageFolder.objects.get(id=folder_id)
        # Check duplicate name within same brand+product
        if ImageFolder.objects.filter(name=new_name, brand=folder.brand, product=folder.product).exclude(id=folder.id).exists():
            return JsonResponse({'success': False, 'message': f'Thư mục "{new_name}" đã tồn tại!'})
        folder.name = new_name
        folder.slug = ''  # reset slug to auto-generate
        folder.save()
        return JsonResponse({'success': True, 'message': f'Đã đổi tên thành "{new_name}"!'})
    except ImageFolder.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Thư mục không tồn tại!'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@login_required
@require_http_methods(["POST"])
@csrf_exempt
def image_folder_delete(request):
    """Xóa thư mục ảnh (và toàn bộ FolderColorImage bên trong)"""
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'message': 'Không có quyền!'}, status=403)

    from store.models import ImageFolder

    folder_id = request.POST.get('folder_id', '').strip()
    if not folder_id:
        return JsonResponse({'success': False, 'message': 'Thiếu folder_id!'}, status=400)

    try:
        folder = ImageFolder.objects.get(id=folder_id)
        folder_name = folder.name
        folder.delete()  # cascade xóa FolderColorImage
        return JsonResponse({'success': True, 'message': f'Đã xóa thư mục "{folder_name}"!'})
    except ImageFolder.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Thư mục không tồn tại!'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)



@login_required
@require_http_methods(["GET"])
def folder_color_image_list(request):
    """Danh sách ảnh theo thư mục + SKU + màu (để hiển thị khi bấm Quản lý)"""
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'message': 'Không có quyền!'}, status=403)

    from store.models import FolderColorImage

    folder_id = request.GET.get('folder_id')
    sku = request.GET.get('sku', '').strip()
    color_name = request.GET.get('color_name', '').strip()

    if not folder_id or not sku or not color_name:
        return JsonResponse({'success': True, 'images': []})

    try:
        imgs = FolderColorImage.objects.filter(
            folder_id=folder_id,
            sku=sku,
            color_name=color_name
        ).order_by('order')

        images = [{'id': img.id, 'url': img.image.url, 'order': img.order} for img in imgs]
        return JsonResponse({'success': True, 'images': images})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)



@login_required
@require_http_methods(["POST"])
@csrf_exempt
def folder_color_image_upload(request):
    """
    Upload ảnh cho 1 màu + SKU trong thư mục:
    - folder_id
    - brand_id (optional)
    - sku
    - color_name
    - image (file)
    """
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'message': 'Không có quyền!'}, status=403)

    from store.models import ImageFolder, FolderColorImage, Brand

    folder_id = request.POST.get('folder_id')
    brand_id = request.POST.get('brand_id')
    sku = request.POST.get('sku', '').strip()
    color_name = request.POST.get('color_name', '').strip()
    image_file = request.FILES.get('image')

    if not folder_id or not sku or not color_name or not image_file:
        return JsonResponse({'success': False, 'message': 'Thiếu thông tin bắt buộc!'}, status=400)

    try:
        folder = ImageFolder.objects.get(id=folder_id)
    except ImageFolder.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Thư mục không tồn tại!'}, status=404)

    brand = None
    if brand_id:
        try:
            brand = Brand.objects.get(id=brand_id)
        except Brand.DoesNotExist:
            brand = None

    # Lấy thứ tự tiếp theo trong nhóm này
    max_order = FolderColorImage.objects.filter(
        folder=folder,
        sku=sku,
        color_name=color_name
    ).aggregate(Max('order'))['order__max'] or 0

    img = FolderColorImage.objects.create(
        folder=folder,
        brand=brand,
        sku=sku,
        color_name=color_name,
        image=image_file,
        order=max_order + 1,
    )

    return JsonResponse({
        'success': True,
        'message': f'Đã upload ảnh màu \"{color_name}\"!',
        'image': {
            'id': img.id,
            'url': img.image.url,
            'order': img.order,
            'folder_name': folder.name,
            'color_name': img.color_name,
            'sku': img.sku,
        }
    })



@login_required
@require_http_methods(["POST"])
@csrf_exempt
def folder_color_image_delete(request):
    """Xóa 1 ảnh màu trong thư mục"""
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'message': 'Không có quyền!'}, status=403)

    from store.models import FolderColorImage

    image_id = request.POST.get('image_id')
    if not image_id:
        return JsonResponse({'success': False, 'message': 'Thiếu thông tin!'}, status=400)

    try:
        img = FolderColorImage.objects.get(id=image_id)
    except FolderColorImage.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Ảnh không tồn tại!'}, status=404)

    # Xóa file vật lý
    try:
        if img.image and hasattr(img.image, 'path'):
            if os.path.isfile(img.image.path):
                os.remove(img.image.path)
            img.image.delete(save=False)
    except Exception:
        pass

    img.delete()
    return JsonResponse({'success': True, 'message': 'Đã xóa ảnh màu!'})



@login_required
@require_http_methods(["POST"])
@csrf_exempt
def folder_color_rename(request):
    """Đổi tên màu cho tất cả ảnh cùng folder + sku + color cũ."""
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'message': 'Không có quyền!'}, status=403)

    from store.models import FolderColorImage

    folder_id = request.POST.get('folder_id')
    sku = request.POST.get('sku', '').strip()
    old_color = request.POST.get('old_color_name', '').strip()
    new_color = request.POST.get('new_color_name', '').strip()

    if not folder_id or not sku or not old_color or not new_color:
        return JsonResponse({'success': False, 'message': 'Thiếu thông tin!'}, status=400)

    if old_color == new_color:
        return JsonResponse({'success': True, 'message': 'Tên màu không thay đổi.'})

    updated = FolderColorImage.objects.filter(
        folder_id=folder_id,
        sku=sku,
        color_name=old_color
    ).update(color_name=new_color)

    if updated == 0:
        return JsonResponse({'success': False, 'message': 'Không tìm thấy ảnh để cập nhật!'})

    return JsonResponse({
        'success': True,
        'message': f'Đã đổi tên màu "{old_color}" → "{new_color}" ({updated} ảnh).'
    })



@login_required
@require_http_methods(["POST"])
@csrf_exempt
def folder_color_row_delete(request):
    """Xóa tất cả ảnh màu theo folder + sku + color (xóa cả row)"""
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'message': 'Không có quyền!'}, status=403)

    from store.models import FolderColorImage

    folder_id = request.POST.get('folder_id')
    sku = request.POST.get('sku', '').strip()
    color_name = request.POST.get('color_name', '').strip()

    if not folder_id or not sku or not color_name:
        return JsonResponse({'success': False, 'message': 'Thiếu thông tin!'}, status=400)

    try:
        images = FolderColorImage.objects.filter(
            folder_id=folder_id,
            sku=sku,
            color_name=color_name
        )
        
        count = images.count()
        if count == 0:
            return JsonResponse({'success': False, 'message': 'Không tìm thấy ảnh nào!'}, status=404)
        
        for img in images:
            try:
                if img.image and hasattr(img.image, 'path'):
                    if os.path.isfile(img.image.path):
                        os.remove(img.image.path)
                    img.image.delete(save=False)
            except Exception:
                pass
            img.delete()
        
        return JsonResponse({'success': True, 'message': f'Đã xóa {count} ảnh màu!'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)



@login_required
@require_http_methods(["GET"])
def get_product_detail(request):
    """Lấy thông tin chi tiết sản phẩm (AJAX)"""
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'message': 'Không có quyền!'}, status=403)
    
    from store.models import Product, ProductDetail, ProductVariant, FolderColorImage, ProductSpecification
    
    product_id = request.GET.get('product_id')
    
    if not product_id:
        return JsonResponse({'success': False, 'message': 'Thiếu thông tin!'}, status=400)
    
    try:
        product = Product.objects.select_related('detail').get(id=product_id)
    except Product.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Sản phẩm không tồn tại!'}, status=404)
    
    # Ưu tiên lấy từ ProductDetail nếu có
    detail = getattr(product, 'detail', None)
    detail_id = detail.id if detail else None
    
    variants = []
    skus_with_color = []
    if detail:
        # Biến thể chi tiết (màu + dung lượng + giá)
        qs = ProductVariant.objects.filter(detail=detail).order_by('price')
        for v in qs:
            variants.append({
                'id': v.id,
                'color_name': v.color_name,
                'color_hex': v.color_hex,
                'storage': v.storage,
                'original_price': int(v.original_price),
                'discount_percent': int(v.discount_percent),
                'price': int(v.price),
                'sku': v.sku,
                'stock_quantity': v.stock_quantity,
            })
    
    # Dropdown "Chọn màu (SKU)": CHỈ lấy SKU của sản phẩm hiện tại từ ProductDetail.sku
    # Không lấy từ FolderColorImage của toàn brand để tránh hiển thị SKU sản phẩm khác
    skus_with_color = []
    
    # Lấy mapping SKU -> màu từ FolderColorImage (để lấy tên màu nếu có)
    sku_to_color = {}
    if product.brand_id:
        rows = FolderColorImage.objects.filter(brand_id=product.brand_id).order_by('sku', 'created_at')
        for row in rows:
            if row.sku and row.sku not in sku_to_color:
                sku_to_color[row.sku] = (row.color_name or '').strip()
    
    # CHỈ lấy SKU đã được thêm cho sản phẩm này trong ProductDetail.sku
    if detail and (detail.sku or '').strip():
        sku_list_raw = [x.strip() for x in (detail.sku or '').split(',') if x.strip()]
        for sku in sku_list_raw:
            color_name = sku_to_color.get(sku, '')
            skus_with_color.append({'sku': sku, 'color_name': color_name})
    
    # Lấy thông số kỹ thuật nếu có
    spec_data = None
    if detail:
        try:
            spec = ProductSpecification.objects.get(detail=detail)
            spec_data = spec.spec_json
        except ProductSpecification.DoesNotExist:
            pass
    
    # Lấy YouTube ID nếu có
    youtube_id = ''
    if detail and detail.youtube_id:
        youtube_id = detail.youtube_id
    
    return JsonResponse({
        'success': True,
        'product_name': product.name,
        'product_image': product.image.url if product.image else None,
        'product_stock': product.stock,
        'detail_id': detail_id,
        'variants': variants,
        'skus_with_color': skus_with_color,
        'spec_data': spec_data,
        'youtube_id': youtube_id,
    })



@login_required
@require_http_methods(["POST"])
@csrf_exempt
def save_youtube_id(request):
    """Lưu YouTube Video ID cho sản phẩm"""
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'message': 'Không có quyền!'}, status=403)
    
    from store.models import Product, ProductDetail
    
    product_id = request.POST.get('product_id')
    youtube_id = request.POST.get('youtube_id', '').strip()
    
    if not product_id:
        return JsonResponse({'success': False, 'message': 'Thiếu thông tin sản phẩm!'}, status=400)
    
    try:
        product = Product.objects.get(id=product_id)
    except Product.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Sản phẩm không tồn tại!'}, status=404)
    
    detail, created = ProductDetail.objects.get_or_create(product=product)
    detail.youtube_id = youtube_id
    detail.save(update_fields=['youtube_id'])
    
    if youtube_id:
        return JsonResponse({'success': True, 'message': f'Đã lưu YouTube ID: {youtube_id}', 'youtube_id': youtube_id})
    else:
        return JsonResponse({'success': True, 'message': 'Đã xóa YouTube ID!', 'youtube_id': ''})

@login_required
@require_http_methods(["GET"])
def sku_list(request):
    """Lấy danh sách tất cả SKU"""
    from store.models import ProductDetail
    try:
        # Get all ProductDetails with SKU
        details = ProductDetail.objects.filter(
            sku__isnull=False, 
            sku__gt=''
        ).select_related('product__brand').order_by('-created_at')
        
        skus = []
        for detail in details:
            sku_list = detail.sku.split(',')
            for sku_item in sku_list:
                sku_item = sku_item.strip()
                if sku_item:
                    skus.append({
                        'id': f'{detail.id}_{sku_item}',  # Composite ID
                        'sku': sku_item,
                        'product_id': detail.product.id,
                        'product_name': detail.product.name,
                        'brand_id': detail.product.brand.id if detail.product.brand else None,
                        'brand_name': detail.product.brand.name if detail.product.brand else None,
                        'created_at': detail.created_at.isoformat() if detail.created_at else None
                    })
        
        return JsonResponse({'success': True, 'skus': skus})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@login_required
@require_http_methods(["POST"])
@csrf_exempt
def sku_add(request):
    """Thêm SKU mới cho sản phẩm"""
    from store.models import Product, ProductDetail
    try:
        product_id = request.POST.get('product_id')
        sku = request.POST.get('sku', '').strip()
        
        if not product_id or not sku:
            return JsonResponse({'success': False, 'message': 'Thiếu thông tin!'}, status=400)
        
        product = Product.objects.get(id=product_id)
        
        # Get or create ProductDetail
        detail, created = ProductDetail.objects.get_or_create(product=product)
        
        # Check for duplicate SKU
        existing_skus = detail.sku.split(',') if detail.sku else []
        if sku in existing_skus:
            return JsonResponse({'success': False, 'message': f'SKU "{sku}" đã tồn tại!'}, status=400)
        
        # Add new SKU
        if detail.sku:
            detail.sku += f',{sku}'
        else:
            detail.sku = sku
        detail.save(update_fields=['sku'])
        
        return JsonResponse({'success': True, 'message': f'Đã thêm SKU: {sku}'})
    except Product.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Sản phẩm không tồn tại!'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@login_required
@require_http_methods(["POST"])
@csrf_exempt
def sku_edit(request):
    """Sửa SKU"""
    from store.models import ProductDetail
    try:
        sku_id = request.POST.get('sku_id')
        new_sku = request.POST.get('sku', '').strip()
        
        if not sku_id or not new_sku:
            return JsonResponse({'success': False, 'message': 'Thiếu thông tin!'}, status=400)
        
        # Parse composite ID - split at first underscore from right (sku might contain underscore)
        # Format: detailID_skuvalue
        last_underscore_idx = sku_id.rfind('_')
        if last_underscore_idx == -1:
            return JsonResponse({'success': False, 'message': 'ID không hợp lệ!'}, status=400)
        
        detail_id = sku_id[:last_underscore_idx]
        old_sku = sku_id[last_underscore_idx + 1:]
        
        detail = ProductDetail.objects.get(id=detail_id)
        
        # Replace SKU
        existing_skus = detail.sku.split(',') if detail.sku else []
        new_skus = []
        for s in existing_skus:
            if s.strip() == old_sku:
                new_skus.append(new_sku)
            else:
                new_skus.append(s)
        
        # Check for duplicate (exclude current one being edited)
        if new_sku in [s.strip() for s in new_skus if s.strip() != old_sku]:
            return JsonResponse({'success': False, 'message': f'SKU "{new_sku}" đã tồn tại!'}, status=400)
        
        detail.sku = ','.join(new_skus)
        detail.save(update_fields=['sku'])
        
        return JsonResponse({'success': True, 'message': f'Đã sửa SKU thành: {new_sku}'})
    except ProductDetail.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'SKU không tồn tại!'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@login_required
@require_http_methods(["POST"])
@csrf_exempt
def sku_delete(request):
    """Xóa SKU"""
    from store.models import ProductDetail
    try:
        sku_id = request.POST.get('sku_id')
        
        if not sku_id:
            return JsonResponse({'success': False, 'message': 'Thiếu thông tin!'}, status=400)
        
        # Parse composite ID - split at first underscore from right
        last_underscore_idx = sku_id.rfind('_')
        if last_underscore_idx == -1:
            return JsonResponse({'success': False, 'message': 'ID không hợp lệ!'}, status=400)
        
        detail_id = sku_id[:last_underscore_idx]
        sku_to_delete = sku_id[last_underscore_idx + 1:]
        
        detail = ProductDetail.objects.get(id=detail_id)
        
        # Remove SKU
        existing_skus = detail.sku.split(',') if detail.sku else []
        existing_skus = [s.strip() for s in existing_skus if s.strip() != sku_to_delete]
        
        detail.sku = ','.join(existing_skus)
        detail.save(update_fields=['sku'])
        
        return JsonResponse({'success': True, 'message': f'Đã xóa SKU: {sku_to_delete}'})
    except ProductDetail.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'SKU không tồn tại!'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)



@login_required
@require_http_methods(["POST"])
@csrf_exempt
def product_specification_upload(request):
    """Upload JSON file cho Thông số kỹ thuật sản phẩm"""
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'message': 'Không có quyền!'}, status=403)
    
    from store.models import ProductDetail, ProductSpecification
    import json
    
    try:
        detail_id = request.POST.get('detail_id')
        json_file = request.FILES.get('json_file')
        
        if not detail_id or not json_file:
            return JsonResponse({'success': False, 'message': 'Thiếu thông tin hoặc file!'}, status=400)
        
        # Validate file type
        if not json_file.name.endswith('.json'):
            return JsonResponse({'success': False, 'message': 'File phải có đuôi .json!'}, status=400)
        
        # Get ProductDetail
        detail = ProductDetail.objects.get(id=detail_id)
        
        # Read and validate JSON
        try:
            file_content = json_file.read().decode('utf-8')
            spec_data = json.loads(file_content)
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'File JSON không hợp lệ!'}, status=400)
        except UnicodeDecodeError:
            return JsonResponse({'success': False, 'message': 'Mã hóa file không hợp lệ! Hãy dùng UTF-8'}, status=400)
        
        # Create or update specification
        spec, created = ProductSpecification.objects.get_or_create(detail=detail)
        spec.spec_json = spec_data
        spec.save()
        
        message = 'Tải file thông số kỹ thuật thành công!'
        return JsonResponse({
            'success': True,
            'message': message,
            'spec_id': spec.id,
            'spec_data': spec_data
        })
    
    except ProductDetail.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Chi tiết sản phẩm không tồn tại!'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Lỗi: {str(e)}'}, status=500)



@login_required
@require_http_methods(["POST"])
@csrf_exempt
def product_specification_delete(request):
    """Xóa Thông số kỹ thuật sản phẩm"""
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'message': 'Không có quyền!'}, status=403)
    
    from store.models import ProductDetail, ProductSpecification
    
    try:
        detail_id = request.POST.get('detail_id')
        if not detail_id:
            return JsonResponse({'success': False, 'message': 'Thiếu thông tin!'}, status=400)
        
        detail = ProductDetail.objects.get(id=detail_id)
        
        try:
            spec = ProductSpecification.objects.get(detail=detail)
            spec.delete()
            return JsonResponse({'success': True, 'message': 'Đã xóa thông số kỹ thuật!'})
        except ProductSpecification.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Chưa có thông số kỹ thuật nào!'}, status=404)
    
    except ProductDetail.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Chi tiết sản phẩm không tồn tại!'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Lỗi: {str(e)}'}, status=500)

def banner_list(request):
    """Lấy danh sách tất cả banner - cho phép truy cập công khai"""
    
    from store.models import Banner
    
    try:
        banners = Banner.objects.all().order_by('banner_id', '-created_at')
        banner_data = []
        for banner in banners:
            banner_data.append({
                'id': banner.id,
                'banner_id': banner.banner_id,
                'image_url': banner.image.url,
                'created_at': banner.created_at.strftime('%Y-%m-%d %H:%M:%S')
            })
        return JsonResponse({'success': True, 'banners': banner_data})
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Lỗi: {str(e)}'}, status=500)



def banner_add(request):
    """Thêm banner mới"""
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'message': 'Không có quyền!'}, status=403)
    
    from store.models import Banner
    
    try:
        banner_id = request.POST.get('banner_id')
        image = request.FILES.get('image')
        
        if not banner_id:
            return JsonResponse({'success': False, 'message': 'Vui lòng nhập ID!'}, status=400)
        
        if not image:
            return JsonResponse({'success': False, 'message': 'Vui lòng chọn ảnh!'}, status=400)
        
        # Tạo banner mới
        banner = Banner.objects.create(
            banner_id=banner_id,
            image=image
        )
        
        return JsonResponse({
            'success': True, 
            'message': f'Đã thêm banner ID {banner_id}!',
            'banner': {
                'id': banner.id,
                'banner_id': banner.banner_id,
                'image_url': banner.image.url
            }
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Lỗi: {str(e)}'}, status=500)



@require_POST
def banner_replace(request):
    """Thay thế ảnh banner có cùng banner_id (đè ảnh cũ)"""
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'message': 'Không có quyền!'}, status=403)
    
    from store.models import Banner
    import os
    
    try:
        banner_id = request.POST.get('banner_id')
        image = request.FILES.get('image')
        
        if not banner_id:
            return JsonResponse({'success': False, 'message': 'Vui lòng nhập ID!'}, status=400)
        
        if not image:
            return JsonResponse({'success': False, 'message': 'Vui lòng chọn ảnh!'}, status=400)
        
        # Tìm banner hiện có theo banner_id
        existing = Banner.objects.filter(banner_id=banner_id).first()
        
        if existing:
            # Xóa file ảnh cũ
            if existing.image:
                old_path = existing.image.path
                if os.path.exists(old_path):
                    try:
                        os.remove(old_path)
                    except Exception:
                        pass
            # Gán ảnh mới và lưu
            existing.image = image
            existing.save()
            banner = existing
        else:
            # Nếu chưa có thì tạo mới
            banner = Banner.objects.create(banner_id=banner_id, image=image)
        
        return JsonResponse({
            'success': True,
            'message': f'Đã thay ảnh banner ID {banner_id}!',
            'banner': {
                'id': banner.id,
                'banner_id': banner.banner_id,
                'image_url': banner.image.url
            }
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Lỗi: {str(e)}'}, status=500)



def banner_delete(request):
    """Xóa banner"""
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'message': 'Không có quyền!'}, status=403)
    
    from store.models import Banner
    import os
    
    try:
        banner_id = request.POST.get('banner_id')
        if not banner_id:
            return JsonResponse({'success': False, 'message': 'Thiếu thông tin!'}, status=400)
        
        # Chuyển đổi sang số nếu là số
        try:
            banner_id = int(banner_id)
            banner = Banner.objects.get(id=banner_id)
        except (ValueError, Banner.DoesNotExist):
            # Thử tìm theo banner_id string
            try:
                banner = Banner.objects.get(banner_id=str(banner_id))
            except Banner.DoesNotExist:
                return JsonResponse({'success': False, 'message': 'Banner không tồn tại!'}, status=404)
        
        banner_id_display = banner.banner_id
        
        # Xóa file ảnh khỏi server
        if banner.image:
            image_path = banner.image.path
            if os.path.exists(image_path):
                try:
                    os.remove(image_path)
                except Exception as e:
                    print(f"Lỗi khi xóa file ảnh: {e}")
        
        banner.delete()
        
        return JsonResponse({'success': True, 'message': f'Đã xóa banner ID {banner_id_display}!'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Lỗi: {str(e)}'}, status=500)

@csrf_exempt
@login_required
def product_content_list(request):
    """Lấy danh sách tất cả nội dung sản phẩm"""
    from store.models import ProductContent
    
    contents = ProductContent.objects.select_related('brand', 'product').all().order_by('-created_at')
    content_data = []
    for content in contents:
        content_data.append({
            'id': content.id,
            'brand_id': content.brand.id,
            'brand_name': content.brand.name,
            'product_id': content.product.id,
            'product_name': content.product.name,
            'content_text': content.content_text,
            'image_url': content.image.url if content.image else None,
            'created_at': content.created_at.strftime('%Y-%m-%d %H:%M:%S')
        })
    
    return JsonResponse({'success': True, 'contents': content_data})



@csrf_exempt
@login_required
def product_content_add(request):
    """Thêm nội dung sản phẩm mới"""
    import os
    from store.models import ProductContent, Brand, Product
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Phương thức không hợp lệ!'}, status=400)
    
    try:
        brand_id = request.POST.get('brand_id')
        product_id = request.POST.get('product_id')
        content_text = request.POST.get('content_text', '')
        image = request.FILES.get('image')
        
        if not brand_id:
            return JsonResponse({'success': False, 'message': 'Vui lòng chọn hãng!'}, status=400)
        
        if not product_id:
            return JsonResponse({'success': False, 'message': 'Vui lòng chọn sản phẩm!'}, status=400)
        
        # Lấy brand và product
        try:
            brand = Brand.objects.get(id=brand_id)
        except Brand.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Hãng không tồn tại!'}, status=400)
        
        try:
            product = Product.objects.get(id=product_id)
        except Product.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Sản phẩm không tồn tại!'}, status=400)
        
        # Kiểm tra đã tồn tại chưa
        existing = ProductContent.objects.filter(brand=brand, product=product).first()
        
        if existing:
            # Cập nhật nội dung
            existing.content_text = content_text
            if image:
                # Xóa ảnh cũ nếu có
                if existing.image:
                    old_image_path = existing.image.path
                    if os.path.exists(old_image_path):
                        os.remove(old_image_path)
                existing.image = image
            existing.save()
            content = existing
            message = f'Đã cập nhật nội dung cho {product.name}!'
        else:
            # Tạo mới
            content = ProductContent.objects.create(
                brand=brand,
                product=product,
                content_text=content_text,
                image=image
            )
            message = f'Đã thêm nội dung cho {product.name}!'
        
        return JsonResponse({
            'success': True,
            'message': message,
            'content': {
                'id': content.id,
                'brand_id': content.brand.id,
                'brand_name': content.brand.name,
                'product_id': content.product.id,
                'product_name': content.product.name,
                'content_text': content.content_text,
                'image_url': content.image.url if content.image else None
            }
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'message': f'Lỗi server: {str(e)}'}, status=500)



@csrf_exempt
@login_required
def product_content_replace(request):
    """Thay thế ảnh nội dung sản phẩm"""
    import os
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Phương thức không hợp lệ!'}, status=400)
    
    from store.models import ProductContent
    
    content_id = request.POST.get('content_id')
    image = request.FILES.get('image')
    
    if not content_id:
        return JsonResponse({'success': False, 'message': 'Vui lòng cung cấp ID!'}, status=400)
    
    if not image:
        return JsonResponse({'success': False, 'message': 'Vui lòng chọn ảnh!'}, status=400)
    
    try:
        content_id = int(content_id)
        content = ProductContent.objects.get(id=content_id)
    except (ValueError, ProductContent.DoesNotExist):
        return JsonResponse({'success': False, 'message': 'Nội dung không tồn tại!'}, status=404)
    
    # Xóa ảnh cũ nếu có
    if content.image:
        old_image_path = content.image.path
        if os.path.exists(old_image_path):
            os.remove(old_image_path)
    
    # Cập nhật ảnh mới
    content.image = image
    content.save()
    
    return JsonResponse({
        'success': True,
        'message': f'Đã thay ảnh cho {content.product.name}!',
        'content': {
            'id': content.id,
            'content_text': content.content_text,
            'image_url': content.image.url
        }
    })



@csrf_exempt
@login_required
def product_content_delete(request):
    """Xóa nội dung sản phẩm"""
    import os
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Phương thức không hợp lệ!'}, status=400)
    
    from store.models import ProductContent
    
    content_id = request.POST.get('content_id')
    
    if not content_id:
        return JsonResponse({'success': False, 'message': 'Vui lòng cung cấp ID!'}, status=400)
    
    try:
        content_id = int(content_id)
        content = ProductContent.objects.get(id=content_id)
    except (ValueError, ProductContent.DoesNotExist):
        return JsonResponse({'success': False, 'message': 'Nội dung không tồn tại!'}, status=404)
    
    product_name = content.product.name
    
    # Xóa ảnh nếu có
    if content.image:
        image_path = content.image.path
        if os.path.exists(image_path):
            os.remove(image_path)
    
    content.delete()
    
    return JsonResponse({'success': True, 'message': f'Đã xóa nội dung của {product_name}!'})

@login_required
def upload_temp_image(request):
    """Upload ảnh tạm để chèn vào nội dung"""
    import os
    from django.conf import settings
    
    if request.method != 'POST':
        return JsonResponse({'error': {'message': 'Phương thức không hợp lệ!'}}, status=400)
    
    image = request.FILES.get('image') or request.FILES.get('upload')
    
    if not image:
        return JsonResponse({'error': {'message': 'Vui lòng chọn ảnh!'}}, status=400)
    
    # Validate image
    if not image.content_type.startswith('image/'):
        return JsonResponse({'error': {'message': 'File phải là ảnh!'}}, status=400)
    
    # Create temp directory if not exists
    temp_dir = os.path.join(settings.MEDIA_ROOT, 'temp')
    os.makedirs(temp_dir, exist_ok=True)
    
    # Save temp file
    from django.utils import timezone
    import uuid
    ext = os.path.splitext(image.name)[1]
    filename = f"{timezone.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:8]}{ext}"
    filepath = os.path.join(temp_dir, filename)
    
    with open(filepath, 'wb+') as destination:
        for chunk in image.chunks():
            destination.write(chunk)
    
    # Return URL 
    temp_url = f"/media/temp/{filename}"
    
    return JsonResponse({
        'url': temp_url
    })


@login_required
def admin_order_list(request):
    """
    Admin: lấy danh sách đơn hàng (JSON API)
    """
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'message': 'Không có quyền'}, status=403)
    
    from store.models import Order
    
    orders = Order.objects.select_related('user').prefetch_related('items').order_by('-created_at')
    
    result = []
    for order in orders:
        raw_items = list(order.items.all())
        items_data = []
        if raw_items:
            for item in raw_items:
                c_name = item.color_name or '—'
                if ' - ' in c_name:
                    c_name = c_name.split(' - ', 1)[1]
                items_data.append({
                    'product_name': item.product_name,
                    'quantity': item.quantity,
                    'color_name': c_name,
                    'storage': item.storage or '—',
                    'price': str(item.price),
                })
        else:
            items_data.append({
                'product_name': '—',
                'quantity': 0,
                'color_name': '—',
                'storage': '—',
                'price': str(order.total_amount),
            })
        result.append({
            'id': order.id,
            'order_code': order.order_code,
            'total_amount': str(order.total_amount),
            'status': order.status,
            'status_display': order.get_status_display(),
            'payment_method': order.payment_method,
            'refund_account': order.refund_account or '',
            'refund_bank': order.refund_bank or '',
            'refund_status': order.refund_status or '',
            'created_at': order.created_at.strftime('%d/%m/%Y %H:%M'),
            'items': items_data,
        })
    
    return JsonResponse({'success': True, 'orders': result})



@login_required
def best_sellers_admin(request):
    """
    Admin: Trang xem thống kê sản phẩm bán chạy
    Hiển thị tất cả sản phẩm được mua nhiều nhất với chi tiết
    """
    if not request.user.is_superuser:
        messages.error(request, 'Bạn không có quyền truy cập trang này!')
        return redirect('store:home')
    
    from store.models import OrderItem, Product, Order
    from django.db.models import Sum
    
    # Lấy tất cả sản phẩm bán chạy (không giới hạn top 5)
    # Chỉ tính các đơn hàng đã giao thành công (delivered)
    best_sellers_data = OrderItem.objects.filter(
        order__status='delivered',
        product__is_active=True
    ).values(
        'product__id', 
        'product__name', 
        'product__image', 
        'product__price',
        'product__original_price', 
        'product__discount_percent', 
        'product__stock',
        'product__slug',
        'product__brand__name',
        'product__brand__id'
    ).annotate(
        total_sold=Sum('quantity')
    ).order_by('-total_sold')

    # Chuyển đổi queryset thành list với thông tin đầy đủ
    best_sellers = []
    total_sold_all = 0
    for item in best_sellers_data:
        product = Product.objects.filter(id=item['product__id']).first()
        if product:
            best_sellers.append({
                'product': product,
                'total_sold': item['total_sold'],
                'brand_name': item['product__brand__name'],
                'brand_id': item['product__brand__id']
            })
            total_sold_all += item['total_sold']

    context = {
        'best_sellers': best_sellers,
        'total_sold_all': total_sold_all,
    }
    return render(request, 'store/admin/best_sellers_admin.html', context)



@login_required
def best_sellers_api(request):
    """
    API: Lấy danh sách sản phẩm bán chạy (JSON)
    """
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'message': 'Không có quyền'}, status=403)
    
    from store.models import OrderItem, Product
    from django.db.models import Sum
    
    best_sellers_data = OrderItem.objects.filter(
        order__status='delivered',
        product__is_active=True
    ).values(
        'product__id', 
        'product__name', 
        'product__image', 
        'product__price',
        'product__original_price', 
        'product__discount_percent', 
        'product__stock',
        'product__slug',
        'product__brand__name'
    ).annotate(
        total_sold=Sum('quantity')
    ).order_by('-total_sold')[:50]

    items = []
    for item in best_sellers_data:
        product = Product.objects.filter(id=item['product__id']).first()
        if product:
            # Lấy giá
            if product.detail:
                original_price = product.detail.summary_original_price or 0
                discounted_price = product.detail.discounted_price or 0
                discount_percent = product.detail.summary_discount_percent or 0
            else:
                original_price = int(product.original_price or 0) if product.original_price else 0
                discounted_price = int(product.price) if product.price else 0
                discount_percent = product.discount_percent or 0
            
            items.append({
                'id': product.id,
                'name': product.name,
                'brand': item['product__brand__name'] or '',
                'image': product.image.url if product.image else '',
                'original_price': original_price,
                'discounted_price': discounted_price,
                'discount_percent': discount_percent,
                'stock': product.stock,
                'total_sold': item['total_sold'],
                'slug': product.slug,
            })

    return JsonResponse({'success': True, 'items': items})



@login_required
def admin_order_detail(request):
    """
    Admin: chi tiết đơn hàng (JSON API)
    """
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'message': 'Không có quyền'}, status=403)
    
    from store.models import Order, Address
    
    order_id = request.GET.get('id')
    if not order_id:
        return JsonResponse({'success': False, 'message': 'Thiếu ID'}, status=400)
    
    try:
        order = Order.objects.select_related('user').prefetch_related('items').get(id=order_id)
    except Order.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Không tìm thấy đơn hàng'}, status=404)
    
    # Lấy địa chỉ mặc định của người đặt
    address_data = None
    try:
        addr = Address.objects.filter(user=order.user, is_default=True).first()
        if not addr:
            addr = Address.objects.filter(user=order.user).first()
        if addr:
            address_data = {
                'full_name': addr.full_name,
                'phone': addr.phone,
                'address': f"{addr.detail}, {addr.ward_name}, {addr.district_name}, {addr.province_name}",
            }
    except Exception:
        pass
    
    # Items
    items = []
    for item in order.items.all():
        color = item.color_name or ''
        if ' - ' in color:
            color = color.split(' - ', 1)[1]
        items.append({
            'product_name': item.product_name,
            'quantity': item.quantity,
            'color_name': color,
            'storage': item.storage,
            'price': str(item.price),
            'thumbnail': item.thumbnail,
        })
    
    data = {
        'id': order.id,
        'order_code': order.order_code,
        'status': order.status,
        'status_display': order.get_status_display(),
        'payment_method': order.get_payment_method_display(),
        'payment_method_key': order.payment_method,
        'created_at': order.created_at.strftime('%d/%m/%Y %H:%M'),
        'total_amount': str(order.total_amount),
        'user_email': order.user.email if order.user else '—',
        'address': address_data,
        'items': items,
        'voucher': order.coupon_code if order.coupon_code else None,
        'discount_amount': str(int(order.discount_amount)) if order.discount_amount else '0',
        'refund_account': order.refund_account or '',
        'refund_bank': order.refund_bank or '',
        'refund_status': order.refund_status or '',
    }
    
    return JsonResponse({'success': True, 'order': data})



@login_required
@require_POST
def admin_order_update_status(request):
    """
    Admin: cập nhật trạng thái đơn hàng hoặc trạng thái hoàn tiền
    Khi chuyển sang 'delivered' → giảm số lượng tồn kho (HangingProduct + ProductVariant)
    """
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'message': 'Không có quyền'}, status=403)
    
    import json
    from store.models import Order, ProductDetail, HangingProduct
    
    try:
        data = json.loads(request.body)
        order_id = data.get('id')
        new_status = data.get('status')
        refund_status = data.get('refund_status')
        
        order = Order.objects.get(id=order_id)
        old_status = order.status
        
        # Cập nhật trạng thái đơn hàng
        if new_status:
            valid_statuses = ['pending', 'processing', 'shipped', 'delivered', 'cancelled']
            if new_status not in valid_statuses:
                return JsonResponse({'success': False, 'message': 'Trạng thái không hợp lệ'})
            
            # Khi chuyển sang 'delivered' và trước đó chưa phải delivered → giảm stock
            if new_status == 'delivered' and old_status != 'delivered':
                for item in order.items.all():
                    if item.product:
                        # 1. Giảm stock Product.stock (hiển thị trong dashboard Treo sản phẩm)
                        try:
                            product = item.product
                            if product.stock >= item.quantity:
                                product.stock -= item.quantity
                                product.save(update_fields=['stock'])
                        except Exception:
                            pass
                        
                        # 2. Giảm stock HangingProduct (nếu có liên kết)
                        try:
                            hanging = HangingProduct.objects.filter(product=item.product).first()
                            if hanging and hanging.stock_quantity >= item.quantity:
                                hanging.stock_quantity -= item.quantity
                                hanging.save(update_fields=['stock_quantity'])
                        except Exception:
                            pass
                        
                        # 3. Giảm stock ProductVariant (match theo color + storage)
                        try:
                            detail = ProductDetail.objects.get(product=item.product)
                            variant = detail.variants.filter(
                                color_name=item.color_name,
                                storage=item.storage
                            ).first()
                            if variant and variant.stock_quantity >= item.quantity:
                                variant.stock_quantity -= item.quantity
                                variant.save(update_fields=['stock_quantity'])
                        except ProductDetail.DoesNotExist:
                            pass
            
            order.status = new_status
        
        # Cập nhật trạng thái hoàn tiền
        if refund_status is not None:
            valid_refund_statuses = ['', 'pending', 'completed']
            if refund_status not in valid_refund_statuses:
                return JsonResponse({'success': False, 'message': 'Trạng thái hoàn tiền không hợp lệ'})
            order.refund_status = refund_status
        
        order.save()
        
        return JsonResponse({'success': True, 'message': f'Đã cập nhật đơn {order.order_code}'})

    except Order.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Không tìm thấy đơn hàng'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


# ==================== Review Management (Admin) ====================

@csrf_exempt
@login_required
def review_list(request):
    """Lấy danh sách đánh giá (phân trang, tìm kiếm) - chỉ admin"""
    from django.http import JsonResponse
    from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
    from store.models import ProductReview
    from django.db import models

    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'message': 'Bạn không có quyền truy cập!'}, status=403)

    # Lấy tham số
    page = int(request.GET.get('page', 1))
    search = request.GET.get('search', '').strip()
    per_page = 15

    # Query base
    reviews = ProductReview.objects.select_related('user', 'product').order_by('-created_at')

    # Tìm kiếm theo tên sản phẩm hoặc email người dùng
    if search:
        reviews = reviews.filter(
            models.Q(product__name__icontains=search) |
            models.Q(user__email__icontains=search) |
            models.Q(user__last_name__icontains=search)
        )

    # Phân trang
    paginator = Paginator(reviews, per_page)
    try:
        reviews_page = paginator.page(page)
    except PageNotAnInteger:
        reviews_page = paginator.page(1)
    except EmptyPage:
        reviews_page = paginator.page(paginator.num_pages)

    # Build response
    data = []
    for idx, r in enumerate(reviews_page.object_list, start=(reviews_page.number - 1) * per_page + 1):
        # Lấy tên người dùng (ưu tiên last_name, fallback email)
        user_name = r.user.last_name or r.user.email.split('@')[0] if r.user.email else 'Ẩn danh'
        # Lấy ảnh (nếu có)
        images = r.images if isinstance(r.images, list) else []
        data.append({
            'stt': idx,
            'id': r.id,
            'product_name': r.product.name,
            'product_id': r.product.id,
            'user_name': user_name,
            'user_email': r.user.email,
            'rating': r.rating,
            'comment': r.comment or '',
            'images': images,
            'created_at': r.created_at.strftime('%d/%m/%Y %H:%M'),
        })

    return JsonResponse({
        'success': True,
        'reviews': data,
        'total': paginator.count,
        'total_pages': paginator.num_pages,
        'current_page': reviews_page.number,
    })


@csrf_exempt
@login_required
def review_delete(request):
    """Xóa đánh giá - chỉ admin"""
    from django.http import JsonResponse
    from store.models import ProductReview

    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'message': 'Bạn không có quyền truy cập!'}, status=403)

    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Phương thức không hợp lệ!'}, status=400)

    review_id = request.POST.get('review_id')

    if not review_id:
        return JsonResponse({'success': False, 'message': 'Vui lòng cung cấp ID đánh giá!'}, status=400)

    try:
        review_id = int(review_id)
        review = ProductReview.objects.get(id=review_id)
    except (ValueError, ProductReview.DoesNotExist):
        return JsonResponse({'success': False, 'message': 'Đánh giá không tồn tại!'}, status=404)

    product_name = review.product.name
    user_email = review.user.email
    review.delete()

    return JsonResponse({
        'success': True,
        'message': f'Đã xóa đánh giá của {user_email} cho sản phẩm {product_name}'
    })


